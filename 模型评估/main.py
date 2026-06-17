import os
import sys
import threading
import queue
import contextlib
import hashlib
import json
import csv
import importlib
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
import h5py
import matplotlib
matplotlib.use("Qt5Agg")
matplotlib.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Arial Unicode MS"]
matplotlib.rcParams["axes.unicode_minus"] = False
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure

from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import (
    QApplication,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QPlainTextEdit,
    QComboBox,
    QCheckBox,
)

try:
    from DEC import SimpleDECGenerator
except ImportError:
    SimpleDECGenerator = None


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
PROJECT_ROOT = os.path.dirname(BASE_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

from model_name_utils import SHAP_MODEL_KEYS, get_model_display_name, get_model_display_names, normalize_model_key

MODEL_FILE_SUFFIXES = (".pth", ".pkl")
FOLD_MODEL_NAMES = [f"model_fold{idx}.pth" for idx in range(1, 6)]
PKL_FOLD_MODEL_NAMES = [f"model_fold{idx}.pkl" for idx in range(1, 6)]
FIVE_FOLD_MODEL_LABEL = "model_fold1-fold5.pth"
PKL_FIVE_FOLD_MODEL_LABEL = "model_fold1-fold5.pkl"
FIVE_FOLD_MODEL_LABELS = {
    FIVE_FOLD_MODEL_LABEL: FOLD_MODEL_NAMES,
    PKL_FIVE_FOLD_MODEL_LABEL: PKL_FOLD_MODEL_NAMES,
}
PREFERRED_MODEL_FILE_NAMES = [
    *FOLD_MODEL_NAMES,
    *PKL_FOLD_MODEL_NAMES,
    "best_model.pth",
    "best_model.pkl",
    "model.pth",
    "model.pkl",
]


def _find_params_dir_under(run_dir):
    if not run_dir or not os.path.isdir(run_dir):
        return ""
    candidates = []
    for root, _, files in os.walk(run_dir):
        if "params.json" in files:
            candidates.append(root)
    if not candidates:
        return ""
    candidates.sort(key=lambda path: (len(path), path.lower()))
    return candidates[0]


def _file_md5(path):
    digest = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _find_repeat_params_dir_by_model_hash(model_path):
    model_dir = os.path.dirname(model_path) or "."
    try:
        target_hash = _file_md5(model_path)
    except Exception:
        return ""

    try:
        run_names = sorted(os.listdir(model_dir))
    except OSError:
        return ""

    candidate_names = (
        "model.pth",
        "best_model.pth",
        "model.pkl",
        "best_model.pkl",
        *FOLD_MODEL_NAMES,
        *PKL_FOLD_MODEL_NAMES,
    )
    for run_name in run_names:
        run_dir = os.path.join(model_dir, run_name)
        if not run_name.lower().startswith("run_") or not os.path.isdir(run_dir):
            continue
        for root, _, files in os.walk(run_dir):
            file_set = set(files)
            for candidate_name in candidate_names:
                if candidate_name not in file_set:
                    continue
                candidate_path = os.path.join(root, candidate_name)
                try:
                    if _file_md5(candidate_path) == target_hash:
                        return root if os.path.exists(os.path.join(root, "params.json")) else _find_params_dir_under(run_dir)
                except Exception:
                    continue
    return ""


def _resolve_repeat_export_model_dir(model_path):
    """Map repeat-summary exports like CNN/model1.pth back to run_001/.../params.json."""
    model_dir = os.path.dirname(model_path) or "."
    stem = os.path.splitext(os.path.basename(model_path))[0]
    lower_stem = stem.lower()

    if lower_stem.startswith("model"):
        suffix = stem[5:]
        if suffix.isdigit():
            run_dir = os.path.join(model_dir, f"run_{int(suffix):03d}")
            match = _find_params_dir_under(run_dir)
            if match:
                return match

    return _find_repeat_params_dir_by_model_hash(model_path)


def _load_training_metadata(model_path):
    model_dir = os.path.dirname(model_path) or "."
    params_path = os.path.join(model_dir, "params.json")
    if not os.path.exists(params_path):
        repeat_model_dir = _resolve_repeat_export_model_dir(model_path)
        if repeat_model_dir:
            model_dir = repeat_model_dir
            params_path = os.path.join(model_dir, "params.json")
    if not os.path.exists(params_path):
        return model_dir, {}
    with open(params_path, "r", encoding="utf-8") as f:
        return model_dir, json.load(f)


def _existing_path(path_value):
    path_text = str(path_value or "").strip()
    if path_text and os.path.exists(path_text):
        return path_text
    return ""


def _first_existing_path(*path_values):
    for path_value in path_values:
        path_text = _existing_path(path_value)
        if path_text:
            return path_text
    return ""


def _optional_positive_int(text, field_name):
    value_text = str(text or "").strip()
    if not value_text:
        return None
    try:
        value = int(value_text)
    except ValueError as exc:
        raise ValueError(f"{field_name}需要是正整数") from exc
    if value <= 0:
        raise ValueError(f"{field_name}需要是正整数")
    return value


def _load_prediction_runtime():
    prediction_dir = os.path.join(PROJECT_ROOT, "模型预测")
    common_dir = os.path.join(PROJECT_ROOT, "common")
    for path in (prediction_dir, common_dir, PROJECT_ROOT):
        if path and path not in sys.path:
            sys.path.insert(0, path)
    if "predict" in sys.modules:
        loaded_path = getattr(sys.modules["predict"], "__file__", "") or ""
        if os.path.abspath(loaded_path) != os.path.abspath(os.path.join(prediction_dir, "predict.py")):
            del sys.modules["predict"]
    if "model" in sys.modules:
        loaded_path = getattr(sys.modules["model"], "__file__", "") or ""
        if os.path.abspath(os.path.dirname(loaded_path)) != os.path.abspath(prediction_dir):
            del sys.modules["model"]
    return importlib.import_module("predict")


PROBABILITY_MODE_CENTER = "center"
PROBABILITY_MODE_WINDOW_AVERAGE = "window_average"
PROBABILITY_MODE_CHOICES = {
    "中心点原值": PROBABILITY_MODE_CENTER,
    "窗口平均融合": PROBABILITY_MODE_WINDOW_AVERAGE,
}


def _probability_mode_label(mode):
    for label, value in PROBABILITY_MODE_CHOICES.items():
        if value == mode:
            return label
    return "中心点原值"


def _confidence_grid_from_positions(positions, confidences):
    coords = np.asarray(positions, dtype=np.float64)
    values = np.asarray(confidences, dtype=np.float64).reshape(-1)
    if coords.ndim != 2 or coords.shape[1] < 2 or len(coords) != len(values):
        return None

    rounded_x = np.round(coords[:, 0], 8)
    rounded_y = np.round(coords[:, 1], 8)
    x_keys, x_inverse = np.unique(rounded_x, return_inverse=True)
    y_keys, y_inverse = np.unique(rounded_y, return_inverse=True)
    if len(x_keys) * len(y_keys) != len(coords):
        return None

    x_values = np.asarray([coords[rounded_x == key, 0].mean() for key in x_keys], dtype=np.float64)
    y_values = np.asarray([coords[rounded_y == key, 1].mean() for key in y_keys], dtype=np.float64)
    x_order = np.argsort(x_values)
    y_order = np.argsort(y_values)
    x_rank = np.empty_like(x_order)
    y_rank = np.empty_like(y_order)
    x_rank[x_order] = np.arange(len(x_order))
    y_rank[y_order] = np.arange(len(y_order))

    grid = np.full((len(y_values), len(x_values)), np.nan, dtype=np.float64)
    for idx, confidence in enumerate(values):
        row = y_rank[y_inverse[idx]]
        col = x_rank[x_inverse[idx]]
        if np.isfinite(grid[row, col]):
            return None
        grid[row, col] = confidence
    if np.isnan(grid).any():
        return None
    return x_values[x_order], y_values[y_order], grid


def _window_average_grid(grid, metadata):
    meta = metadata or {}
    try:
        window_width = int(meta.get("window_width", meta.get("patch_size", 1)) or 1)
        window_height = int(meta.get("window_height", meta.get("patch_size", window_width)) or window_width)
    except (TypeError, ValueError):
        return None
    if window_width <= 1 and window_height <= 1:
        return None

    values = np.asarray(grid, dtype=np.float64)
    ny, nx = values.shape
    if nx == 0 or ny == 0:
        return None

    left = max(window_width // 2, 0)
    right = max(window_width - left, 1)
    lower = max(window_height // 2, 0)
    upper = max(window_height - lower, 1)
    fused = np.zeros_like(values, dtype=np.float64)
    counts = np.zeros_like(values, dtype=np.float64)

    for row in range(ny):
        row_start = max(0, row - lower)
        row_end = min(ny, row + upper)
        for col in range(nx):
            score = values[row, col]
            if not np.isfinite(score):
                continue
            col_start = max(0, col - left)
            col_end = min(nx, col + right)
            fused[row_start:row_end, col_start:col_end] += score
            counts[row_start:row_end, col_start:col_end] += 1.0

    valid = counts > 0
    fused[valid] /= counts[valid]
    fused[~valid] = np.nan
    return np.clip(fused, 0.0, 1.0)


def _grid_values_for_positions(positions, x_values, y_values, grid):
    coords = np.asarray(positions, dtype=np.float64)
    result = np.full(len(coords), np.nan, dtype=np.float64)
    x_lookup = {round(float(value), 8): idx for idx, value in enumerate(x_values)}
    y_lookup = {round(float(value), 8): idx for idx, value in enumerate(y_values)}
    for idx, pos in enumerate(coords):
        col = x_lookup.get(round(float(pos[0]), 8))
        row = y_lookup.get(round(float(pos[1]), 8))
        if row is not None and col is not None:
            result[idx] = grid[row, col]
    return result


def apply_probability_mode(prediction_data, mode, log_callback=None, context="评估"):
    mode = mode or PROBABILITY_MODE_CENTER
    if mode == PROBABILITY_MODE_CENTER:
        return prediction_data
    if mode != PROBABILITY_MODE_WINDOW_AVERAGE:
        return prediction_data

    log = log_callback or (lambda _message: None)
    positions = prediction_data.get("positions")
    confidences = prediction_data.get("confidences")
    metadata = prediction_data.get("metadata") or {}
    rebuilt = _confidence_grid_from_positions(positions, confidences)
    if rebuilt is None:
        log(f"[{context}] 无法从预测点重建规则网格，已回退为中心点原值概率。")
        return prediction_data

    x_values, y_values, grid = rebuilt
    fused_grid = _window_average_grid(grid, metadata)
    if fused_grid is None:
        log(f"[{context}] 预测文件缺少有效窗口大小信息，已回退为中心点原值概率。")
        return prediction_data

    fused_values = _grid_values_for_positions(positions, x_values, y_values, fused_grid)
    original_values = np.asarray(confidences, dtype=np.float64).reshape(-1)
    invalid = ~np.isfinite(fused_values)
    if np.any(invalid):
        fused_values[invalid] = original_values[invalid]
        log(f"[{context}] 部分融合概率为空，已对 {int(np.sum(invalid))} 个点保留中心点原值。")

    result = dict(prediction_data)
    result["confidences"] = fused_values.astype(np.float32)
    result["predictions"] = None
    result["probability_mode"] = mode
    log(f"[{context}] 已使用窗口平均融合概率参与计算。")
    return result


def build_prediction_data_from_model(config, log_callback):
    """Run prediction in-memory from a trained model file and return evaluation data."""
    predict_runtime = _load_prediction_runtime()
    import torch
    from torch.utils.data import DataLoader

    model_path = config["model_path"]
    model_dir, training_meta = _load_training_metadata(model_path)
    model_type = normalize_model_key(training_meta.get("model"))
    if not model_type:
        model_type = normalize_model_key(Path(model_dir).name.split("_")[0])
    if not model_type:
        raise ValueError("无法确定模型类型，请确认模型目录中存在 params.json。")

    data_path = _first_existing_path(config.get("data_path"), training_meta.get("dataset"))
    label_path = _first_existing_path(
        config.get("label_file"),
        config.get("deposit_file"),
        training_meta.get("test_mineral_path"),
        training_meta.get("label_path"),
        os.path.join(model_dir, "test_minerals.txt"),
    )
    if not data_path:
        raise ValueError("缺少完整特征 H5 文件，请选择完整 H5 或确认训练目录含 params.json。")
    if not label_path:
        raise ValueError("缺少矿点或标签文件。")
    if not os.path.exists(data_path):
        raise FileNotFoundError(f"完整特征 H5 不存在: {data_path}")
    if not os.path.exists(label_path):
        raise FileNotFoundError(f"矿点或标签文件不存在: {label_path}")

    device_pref = str(training_meta.get("device") or "cuda").lower()
    device = torch.device("cuda" if device_pref == "cuda" and torch.cuda.is_available() else "cpu")
    log_callback(f"直接评估模型: {model_path}")
    log_callback(f"模型类型: {get_model_display_name(model_type)}，设备: {device}")
    log_callback(f"完整H5: {data_path}")
    log_callback(f"标签/矿点文件: {label_path}")

    prior = 0.2
    if training_meta.get("resolved_prior") is not None:
        prior = float(training_meta["resolved_prior"])
    elif training_meta.get("calculated_prior") is not None:
        prior = float(training_meta["calculated_prior"])
    elif training_meta.get("manual_prior") is not None and not training_meta.get("auto_prior", False):
        prior = float(training_meta["manual_prior"])

    norm_params_path = _first_existing_path(
        config.get("normalization_params_path"),
        training_meta.get("normalization_params_path"),
        os.path.join(model_dir, "normalization_params.pth"),
    )

    patch_size = int(training_meta.get("patch_size") or training_meta.get("img_size") or 16)
    training_patch_stride = int(training_meta.get("patch_stride") or patch_size)
    prediction_patch_stride = int(config.get("prediction_patch_stride") or config.get("patch_stride") or training_patch_stride)
    log_callback(f"训练窗口大小: {patch_size} | 训练步长: {training_patch_stride} | 本次评估步长: {prediction_patch_stride}")

    dataset = predict_runtime.PredictionDataset(
        data_path,
        label_path,
        model_dir,
        use_custom_norm=True,
        norm_params_path=norm_params_path,
        patch_size=patch_size,
        patch_stride=prediction_patch_stride,
        use_reflect_padding=bool(training_meta.get("reflect_padding", True)),
        selected_channels=predict_runtime.parse_selected_channels(training_meta.get("selected_channels")),
    )
    if getattr(dataset, "feature_mode", "") == "windows":
        log_callback("当前H5已包含预切好的 windows，评估步长不会重新切割该文件。")
    if len(dataset) == 0:
        raise RuntimeError("完整 H5 中没有可预测样本。")

    sample_data = dataset[0][0]
    input_channels, input_height, input_width = sample_data.shape
    input_dim = input_channels * input_height * input_width
    model, _ = predict_runtime.load_model(
        model_type,
        model_path,
        prior=prior,
        input_dim=input_dim,
        device=device,
        data_shape=(input_channels, input_height, input_width),
    )

    batch_size = int(training_meta.get("batchsize") or 32)
    dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=False)
    all_positions = []
    all_predictions = []
    all_confidences = []
    all_labels = []
    log_callback("开始基于模型文件生成评估预测...")
    for batch_idx, (data, positions, labels) in enumerate(dataloader, start=1):
        predictions, confidences = predict_runtime.predict_batch(model, data, device)
        all_positions.extend([pos.detach().cpu().numpy() for pos in positions])
        all_predictions.extend(predictions)
        all_confidences.extend(np.asarray(confidences, dtype=np.float32).reshape(-1).tolist())
        all_labels.append(labels.detach().cpu().numpy())
        if batch_idx % 50 == 0:
            log_callback(f"已预测 {batch_idx} 个批次")

    positions_array = np.vstack(all_positions).astype(np.float32)
    offset_x, offset_y = getattr(dataset, "coord_offset", (0.0, 0.0))
    positions_array[:, 0] += float(offset_x)
    positions_array[:, 1] += float(offset_y)
    predictions_numeric = np.asarray([1 if p == "Positive" else -1 for p in all_predictions], dtype=np.int32)
    labels_array = np.concatenate(all_labels).astype(np.int32)
    labels_array = np.where(labels_array > 0, 1, -1)
    log_callback(f"评估预测完成: {len(positions_array)} 个预测点")

    return {
        "positions": positions_array,
        "confidences": np.asarray(all_confidences, dtype=np.float32),
        "predictions": predictions_numeric,
        "labels": labels_array,
        "metadata": getattr(dataset, "metadata", None),
        "format": "model",
        "path": model_path,
    }


def run_gdr_task(config, log_callback):
    """Compute GDR/SR metrics for an existing prediction file."""
    from GDR_calculator import GDRCalculator

    def log(msg):
        log_callback(msg)

    calculator = GDRCalculator(
        deposit_file=config["deposit_file"],
        distance_threshold=config["distance_threshold"],
        confidence_threshold=config["confidence_threshold"],
    )

    if config.get("model_path"):
        prediction_data = build_prediction_data_from_model(config, log)
    else:
        log("加载预测文件...")
        prediction_data = calculator.load_predictions(
            config["prediction_file"], positions_key="positions", confidences_key="confidences"
        )
    probability_mode = config.get("probability_mode", PROBABILITY_MODE_CENTER)
    log(f"概率计算方式: {_probability_mode_label(probability_mode)}")
    prediction_data = apply_probability_mode(prediction_data, probability_mode, log, "模型评估")

    test_metrics = None
    test_area_mask = None
    if config.get("independent_test"):
        strategy_text = "手动阈值" if config.get("threshold_strategy") == "fixed" else "EI 最大"
        log(f"独立测试集模式：仅使用测试矿点和测试区域，阈值策略为 {strategy_text}。")
        test_area_mask = calculator.load_test_area_mask(prediction_data, config.get("test_area_file"))
        fixed_threshold = None
        if config.get("threshold_strategy") == "fixed":
            fixed_threshold = config["confidence_threshold"]
        test_metrics = calculator.calculate_independent_test_metrics(
            prediction_data,
            test_area_mask,
            threshold_step=config.get("threshold_step", 0.01),
            fixed_threshold=fixed_threshold,
        )
        confidence_threshold = test_metrics["threshold"]
        gdr = test_metrics["sr"]
        hit_deposits = test_metrics["hit_deposits"]
        total_deposits = test_metrics["total_deposits"]
    else:
        log("计算 GDR...")
        confidence_threshold = config["confidence_threshold"]
        gdr, hit_deposits, total_deposits, _ = calculator.calculate_gdr(prediction_data)

    confidences = np.asarray(prediction_data["confidences"], dtype=float)
    predictions = prediction_data.get("predictions")
    if predictions is None or config.get("independent_test") or probability_mode != PROBABILITY_MODE_CENTER:
        predictions = np.where(confidences > confidence_threshold, 1, -1)
    predictions = np.asarray(predictions).astype(int).flatten()
    predictions = np.where(predictions > 0, 1, -1)
    total_preds = len(predictions)
    scoped_predictions = predictions
    if test_area_mask is not None and len(test_area_mask) == total_preds:
        scoped_predictions = predictions[test_area_mask]
    positive_preds = int(np.sum(scoped_predictions == 1))
    pr_positive = positive_preds / len(scoped_predictions) if len(scoped_predictions) else 0.0
    gdr_over_pr = gdr / pr_positive if pr_positive > 0 else 0.0
    if test_metrics is not None:
        par = test_metrics["pa"]
        gdr_over_pr = test_metrics["ei"]
    else:
        par = (
            float(np.sum(confidences > confidence_threshold) / len(confidences))
            if len(confidences) > 0
            else 0.0
        )
    prior_probability = config["prior"]

    labels = None
    if prediction_data.get("labels") is not None:
        labels = prediction_data["labels"]
    elif config.get("prediction_file", "").lower().endswith((".h5", ".hdf5")):
        try:
            with h5py.File(config["prediction_file"], "r") as f:
                if "labels" in f:
                    labels = f["labels"][:]
        except Exception as exc:
            log(f"警告: 无法读取标签信息: {exc}")

    labels_available = labels is not None and len(labels) == total_preds
    if labels is not None and len(labels) == total_preds:
        labels = labels.astype(int).reshape(-1)
        labels = np.where(labels > 0, 1, -1)
        scoped_labels = labels
        scoped_preds_for_confusion = predictions
        if test_area_mask is not None and len(test_area_mask) == total_preds:
            scoped_labels = labels[test_area_mask]
            scoped_preds_for_confusion = predictions[test_area_mask]
        true_positives = int(np.sum((scoped_labels == 1) & (scoped_preds_for_confusion == 1)))
        false_negatives = int(np.sum((scoped_labels == 1) & (scoped_preds_for_confusion == -1)))
        true_negatives = int(np.sum((scoped_labels == -1) & (scoped_preds_for_confusion == -1)))
        false_positives = int(np.sum((scoped_labels == -1) & (scoped_preds_for_confusion == 1)))
        confusion = {
            "tp": true_positives,
            "fn": false_negatives,
            "tn": true_negatives,
            "fp": false_positives,
        }
    else:
        confusion = None
        if not labels_available:
            log("未提供标签信息，跳过混淆矩阵统计。")

    return {
        "gdr": gdr,
        "hit_deposits": hit_deposits,
        "total_deposits": total_deposits,
        "pr_positive": pr_positive,
        "gdr_over_pr": gdr_over_pr,
        "par": par,
        "prior": prior_probability,
        "confusion": confusion,
        "labels_available": labels_available,
        "independent_test": bool(config.get("independent_test")),
        "best_threshold": confidence_threshold,
        "test_metrics": test_metrics,
        "probability_mode": probability_mode,
    }


def run_gdr_batch_task(config, log_callback):
    """Evaluate multiple model files with one shared evaluation configuration."""
    model_entries = list(config.get("model_entries") or [])
    model_paths = list(config.get("model_paths") or [])
    if model_entries:
        model_paths = [str(item.get("path") or "").strip() for item in model_entries if str(item.get("path") or "").strip()]
    if not model_paths:
        single_path = str(config.get("model_path") or "").strip()
        if single_path:
            model_paths = [single_path]
    if not model_paths:
        return run_gdr_task(config, log_callback)

    results = []
    total = len(model_paths)
    for index, model_path in enumerate(model_paths, start=1):
        entry = model_entries[index - 1] if index - 1 < len(model_entries) else {}
        model_name = str(entry.get("display_name") or entry.get("model_name") or Path(str(model_path)).stem)
        log_callback(f"[{index}/{total}] 开始评估 {model_name}")
        item_config = dict(config)
        item_config["model_path"] = str(model_path)
        item_config["model_paths"] = [str(model_path)]
        if config.get("use_model_metadata_per_folder"):
            try:
                model_dir, training_meta = _load_training_metadata(str(model_path))
            except Exception:
                model_dir = os.path.dirname(str(model_path))
                training_meta = {}
            data_path = _first_existing_path(item_config.get("data_path"), training_meta.get("dataset"))
            if data_path:
                item_config["data_path"] = data_path
            norm_params_path = _first_existing_path(
                os.path.join(model_dir, "normalization_params.pth"),
                training_meta.get("normalization_params_path"),
            )
            if norm_params_path:
                item_config["normalization_params_path"] = norm_params_path
            if item_config.get("independent_test"):
                test_area_path = _first_existing_path(
                    os.path.join(model_dir, "test_area.h5"),
                    training_meta.get("test_area_path"),
                    item_config.get("test_area_file"),
                )
                deposit_path = _first_existing_path(
                    os.path.join(model_dir, "test_minerals.txt"),
                    training_meta.get("test_mineral_path"),
                    training_meta.get("label_path"),
                    item_config.get("deposit_file"),
                )
                if test_area_path:
                    item_config["test_area_file"] = test_area_path
            else:
                deposit_path = _first_existing_path(
                    training_meta.get("label_path"),
                    os.path.join(model_dir, "test_minerals.txt"),
                    item_config.get("deposit_file"),
                )
            if deposit_path:
                item_config["deposit_file"] = deposit_path
                item_config["label_file"] = deposit_path
        try:
            metrics = run_gdr_task(item_config, log_callback)
        except Exception as exc:
            log_callback(f"[{index}/{total}] 评估失败: {exc}")
            metrics = {
                "status": "失败",
                "error": str(exc),
                "independent_test": bool(config.get("independent_test")),
                "labels_available": False,
                "confusion": None,
            }
        else:
            metrics["status"] = "成功"
            metrics["error"] = ""
        metrics["model_path"] = str(model_path)
        metrics["model_name"] = model_name
        metrics["folder_name"] = str(entry.get("folder_name") or "")
        metrics["training_dir"] = str(entry.get("training_dir") or Path(str(model_path)).parent.name)
        results.append(metrics)

    aggregate_confusion = None
    confusion_items = [item.get("confusion") for item in results if item.get("confusion")]
    if confusion_items:
        aggregate_confusion = {
            key: int(sum(int(conf.get(key, 0) or 0) for conf in confusion_items))
            for key in ("tp", "fp", "tn", "fn")
        }

    batch_result = {
        "batch": True,
        "models": results,
        "model_names": [item.get("model_name", f"model_{idx}") for idx, item in enumerate(results, start=1)],
        "independent_test": bool(config.get("independent_test")),
        "confusion": aggregate_confusion,
        "labels_available": bool(aggregate_confusion),
    }
    csv_output_path = str(config.get("csv_output_path") or "").strip()
    if csv_output_path:
        save_gdr_batch_csv(batch_result, csv_output_path)
        batch_result["csv_output_path"] = csv_output_path
        log_callback(f"批量评估结果已保存到: {csv_output_path}")
    return batch_result


def save_gdr_batch_csv(metrics: dict, output_path: str):
    """Save GDR batch metrics to CSV."""
    rows = list(metrics.get("models") or [])
    fieldnames = [
        "status",
        "error",
        "folder_name",
        "training_dir",
        "model_name",
        "model_path",
        "mode",
        "threshold",
        "sr_gdr",
        "hit_deposits",
        "total_deposits",
        "pa_par",
        "ei_or_gdr_over_pr",
        "test_area_count",
        "high_potential_count",
        "mean_min_distance",
        "median_min_distance",
        "tp",
        "fp",
        "tn",
        "fn",
    ]
    output = Path(output_path)
    if output.suffix.lower() != ".csv":
        output = output.with_suffix(".csv")
    output.parent.mkdir(parents=True, exist_ok=True)
    with open(output, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for item in rows:
            test_metrics = item.get("test_metrics") or {}
            confusion = item.get("confusion") or {}
            writer.writerow(
                {
                    "status": item.get("status", "成功"),
                    "error": item.get("error", ""),
                    "folder_name": item.get("folder_name", ""),
                    "training_dir": item.get("training_dir", ""),
                    "model_name": item.get("model_name", ""),
                    "model_path": item.get("model_path", ""),
                    "mode": "独立测试集" if item.get("independent_test") else "普通评估",
                    "threshold": item.get("best_threshold", ""),
                    "sr_gdr": item.get("gdr", ""),
                    "hit_deposits": item.get("hit_deposits", ""),
                    "total_deposits": item.get("total_deposits", ""),
                    "pa_par": item.get("par", ""),
                    "ei_or_gdr_over_pr": item.get("gdr_over_pr", ""),
                    "test_area_count": test_metrics.get("test_area_count", ""),
                    "high_potential_count": test_metrics.get("high_potential_count", ""),
                    "mean_min_distance": test_metrics.get("mean_min_distance", ""),
                    "median_min_distance": test_metrics.get("median_min_distance", ""),
                    "tp": confusion.get("tp", ""),
                    "fp": confusion.get("fp", ""),
                    "tn": confusion.get("tn", ""),
                    "fn": confusion.get("fn", ""),
                }
            )


def _finite_numbers(values):
    numbers = []
    for value in values:
        try:
            number = float(value)
        except (TypeError, ValueError):
            continue
        if np.isfinite(number):
            numbers.append(number)
    return numbers


def _mean_std_dict(values):
    numbers = _finite_numbers(values)
    if not numbers:
        return {"mean": "", "std": "", "count": 0}
    return {
        "mean": float(np.mean(numbers)),
        "std": float(np.std(numbers, ddof=1)) if len(numbers) > 1 else 0.0,
        "count": len(numbers),
    }


def _summarize_fold_models(models):
    success_models = [item for item in models if item.get("status", "成功") == "成功"]

    def values_for(key):
        return [item.get(key) for item in success_models]

    hit_ratios = []
    unit_ratios = []
    for item in success_models:
        total_deposits = float(item.get("total_deposits", 0) or 0)
        if total_deposits > 0:
            hit_ratios.append(float(item.get("hit_deposits", 0) or 0) / total_deposits)
        test_metrics = item.get("test_metrics") or {}
        test_area_count = float(test_metrics.get("test_area_count", 0) or 0)
        if test_area_count > 0:
            unit_ratios.append(float(test_metrics.get("high_potential_count", 0) or 0) / test_area_count)

    confusion = None
    confusion_items = [item.get("confusion") for item in success_models if item.get("confusion")]
    if confusion_items:
        confusion = {
            key: int(sum(int(conf.get(key, 0) or 0) for conf in confusion_items))
            for key in ("tp", "fp", "tn", "fn")
        }

    return {
        "fold_count": len(models),
        "success_count": len(success_models),
        "status": "成功" if len(success_models) == len(models) else ("失败" if not success_models else "部分成功"),
        "sr": _mean_std_dict(values_for("gdr")),
        "pa": _mean_std_dict(values_for("par")),
        "ei": _mean_std_dict(values_for("gdr_over_pr")),
        "threshold": _mean_std_dict(values_for("best_threshold")),
        "hit_ratio": _mean_std_dict(hit_ratios),
        "unit_ratio": _mean_std_dict(unit_ratios),
        "confusion": confusion,
    }


def run_gdr_grouped_folds_task(config, log_callback):
    """Evaluate selected model files for every training folder and summarize each folder."""
    model_groups = list(config.get("model_groups") or [])
    group_results = []
    total = len(model_groups)
    for index, group in enumerate(model_groups, start=1):
        entries = list(group.get("entries") or [])
        group_name = str(group.get("display_name") or group.get("folder_name") or f"group_{index}")
        log_callback(f"[{index}/{total}] 开始分组评估 {group_name}")
        group_config = dict(config)
        group_config["model_entries"] = entries
        group_config["model_paths"] = [entry["path"] for entry in entries]
        group_config["model_path"] = entries[0]["path"] if entries else ""
        group_config["csv_output_path"] = ""
        batch_result = run_gdr_batch_task(group_config, log_callback)
        models = list(batch_result.get("models") or [])
        summary = _summarize_fold_models(models)
        group_results.append(
            {
                "display_name": group_name,
                "folder_name": str(group.get("folder_name") or ""),
                "training_dir": str(group.get("training_dir") or ""),
                "training_path": str(group.get("training_path") or ""),
                "models": models,
                "summary": summary,
                "batch_result": batch_result,
            }
        )
        log_callback(
            f"[{index}/{total}] 分组完成 {group_name}: "
            f"有效 {summary['success_count']}/{summary['fold_count']}"
        )

    aggregate_confusion = None
    confusion_items = [
        item.get("summary", {}).get("confusion")
        for item in group_results
        if item.get("summary", {}).get("confusion")
    ]
    if confusion_items:
        aggregate_confusion = {
            key: int(sum(int(conf.get(key, 0) or 0) for conf in confusion_items))
            for key in ("tp", "fp", "tn", "fn")
        }

    result = {
        "grouped_folds": True,
        "groups": group_results,
        "model_names": list(config.get("group_model_names") or FOLD_MODEL_NAMES),
        "independent_test": bool(config.get("independent_test")),
        "confusion": aggregate_confusion,
        "labels_available": bool(aggregate_confusion),
    }
    csv_output_path = str(config.get("csv_output_path") or "").strip()
    if csv_output_path:
        if Path(csv_output_path).suffix.lower() == ".xlsx":
            save_gdr_grouped_folds_xlsx(result, csv_output_path)
        else:
            save_gdr_grouped_folds_csv(result, csv_output_path)
        result["csv_output_path"] = csv_output_path
        log_callback(f"分组批量评估结果已保存到: {csv_output_path}")
    return result


def _format_stat_text(stat, decimals=4):
    stat = stat or {}
    try:
        mean_value = float(stat.get("mean"))
        std_value = float(stat.get("std"))
    except (TypeError, ValueError):
        return ""
    if not np.isfinite(mean_value) or not np.isfinite(std_value):
        return ""
    return f"{mean_value:.{decimals}f} ± {std_value:.{decimals}f}"


def _format_ratio_text(numerator, denominator):
    try:
        numerator = int(numerator)
        denominator = int(denominator)
    except (TypeError, ValueError):
        return ""
    return f"{numerator}/{denominator}"


def save_gdr_grouped_folds_csv(metrics: dict, output_path: str):
    """Save grouped five-fold metrics to CSV, one row per training folder."""
    output = Path(output_path)
    if output.suffix.lower() != ".csv":
        output = output.with_suffix(".csv")
    output.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "folder_name",
        "training_dir",
        "training_path",
        "status",
        "success_count",
        "fold_count",
        "sr_mean",
        "sr_std",
        "pa_mean",
        "pa_std",
        "ei_mean",
        "ei_std",
        "threshold_mean",
        "threshold_std",
        "hit_ratio_mean",
        "hit_ratio_std",
        "unit_ratio_mean",
        "unit_ratio_std",
        "tp",
        "fp",
        "tn",
        "fn",
    ]
    model_names = list(metrics.get("model_names") or FOLD_MODEL_NAMES)
    for model_name in model_names:
        fold_key = Path(model_name).stem
        fieldnames.extend(
            [
                f"{fold_key}_path",
                f"{fold_key}_status",
                f"{fold_key}_sr",
                f"{fold_key}_pa",
                f"{fold_key}_ei",
                f"{fold_key}_threshold",
                f"{fold_key}_error",
            ]
        )

    with open(output, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for group in metrics.get("groups") or []:
            summary = group.get("summary") or {}
            confusion = summary.get("confusion") or {}
            row = {
                "folder_name": group.get("folder_name", ""),
                "training_dir": group.get("training_dir", ""),
                "training_path": group.get("training_path", ""),
                "status": summary.get("status", ""),
                "success_count": summary.get("success_count", ""),
                "fold_count": summary.get("fold_count", ""),
                "sr_mean": (summary.get("sr") or {}).get("mean", ""),
                "sr_std": (summary.get("sr") or {}).get("std", ""),
                "pa_mean": (summary.get("pa") or {}).get("mean", ""),
                "pa_std": (summary.get("pa") or {}).get("std", ""),
                "ei_mean": (summary.get("ei") or {}).get("mean", ""),
                "ei_std": (summary.get("ei") or {}).get("std", ""),
                "threshold_mean": (summary.get("threshold") or {}).get("mean", ""),
                "threshold_std": (summary.get("threshold") or {}).get("std", ""),
                "hit_ratio_mean": (summary.get("hit_ratio") or {}).get("mean", ""),
                "hit_ratio_std": (summary.get("hit_ratio") or {}).get("std", ""),
                "unit_ratio_mean": (summary.get("unit_ratio") or {}).get("mean", ""),
                "unit_ratio_std": (summary.get("unit_ratio") or {}).get("std", ""),
                "tp": confusion.get("tp", ""),
                "fp": confusion.get("fp", ""),
                "tn": confusion.get("tn", ""),
                "fn": confusion.get("fn", ""),
            }
            models_by_name = {Path(str(item.get("model_path", ""))).name: item for item in group.get("models") or []}
            for model_name in model_names:
                fold_key = Path(model_name).stem
                item = models_by_name.get(model_name, {})
                row.update(
                    {
                        f"{fold_key}_path": item.get("model_path", ""),
                        f"{fold_key}_status": item.get("status", ""),
                        f"{fold_key}_sr": item.get("gdr", ""),
                        f"{fold_key}_pa": item.get("par", ""),
                        f"{fold_key}_ei": item.get("gdr_over_pr", ""),
                        f"{fold_key}_threshold": item.get("best_threshold", ""),
                        f"{fold_key}_error": item.get("error", ""),
                    }
                )
            writer.writerow(row)


def save_gdr_grouped_folds_xlsx(metrics: dict, output_path: str):
    """Save grouped five-fold metrics to a workbook with summary and fold details."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output = Path(output_path)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总"
    ws_detail = wb.create_sheet("fold明细")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    summary_headers = [
        "模型目录",
        "状态",
        "有效模型",
        "SR mean ± std",
        "PA mean ± std",
        "EI mean ± std",
        "阈值 mean ± std",
        "命中测试矿点 mean ± std",
        "测试区预测单元 mean ± std",
        "TP",
        "FP",
        "TN",
        "FN",
        "训练目录",
        "完整路径",
    ]
    ws_summary.append(summary_headers)

    for group in metrics.get("groups") or []:
        summary = group.get("summary") or {}
        confusion = summary.get("confusion") or {}
        ws_summary.append(
            [
                group.get("folder_name", ""),
                summary.get("status", ""),
                f"{summary.get('success_count', 0)}/{summary.get('fold_count', 0)}",
                _format_stat_text(summary.get("sr")),
                _format_stat_text(summary.get("pa")),
                _format_stat_text(summary.get("ei")),
                _format_stat_text(summary.get("threshold")),
                _format_stat_text(summary.get("hit_ratio")),
                _format_stat_text(summary.get("unit_ratio")),
                confusion.get("tp", ""),
                confusion.get("fp", ""),
                confusion.get("tn", ""),
                confusion.get("fn", ""),
                group.get("training_dir", ""),
                group.get("training_path", ""),
            ]
        )

    detail_headers = [
        "模型目录",
        "fold",
        "状态",
        "SR",
        "命中测试矿点",
        "PA",
        "测试区预测单元",
        "EI",
        "阈值",
        "TP",
        "FP",
        "TN",
        "FN",
        "模型路径",
        "错误信息",
    ]
    ws_detail.append(detail_headers)

    for group in metrics.get("groups") or []:
        folder_name = group.get("folder_name", "")
        models_by_name = {Path(str(item.get("model_path", ""))).name: item for item in group.get("models") or []}
        for model_name in list(metrics.get("model_names") or FOLD_MODEL_NAMES):
            item = models_by_name.get(model_name, {})
            test_metrics = item.get("test_metrics") or {}
            confusion = item.get("confusion") or {}
            ws_detail.append(
                [
                    folder_name,
                    Path(model_name).stem,
                    item.get("status", ""),
                    item.get("gdr", ""),
                    _format_ratio_text(item.get("hit_deposits"), item.get("total_deposits")),
                    item.get("par", ""),
                    _format_ratio_text(test_metrics.get("high_potential_count"), test_metrics.get("test_area_count")),
                    item.get("gdr_over_pr", ""),
                    item.get("best_threshold", ""),
                    confusion.get("tp", ""),
                    confusion.get("fp", ""),
                    confusion.get("tn", ""),
                    confusion.get("fn", ""),
                    item.get("model_path", ""),
                    item.get("error", ""),
                ]
            )

    for ws in (ws_summary, ws_detail):
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = left if isinstance(cell.value, str) and len(cell.value) > 20 else center
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            width = min(max(max_len + 2, 10), 55)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output)


def run_dec_task(models: list[dict], log_callback, deposit_file: str, config: dict | None = None):
    from GDR_calculator import GDRCalculator
    config = dict(config or {})
    use_test_area = bool(config.get("use_test_area"))
    test_area_file = str(config.get("test_area_file") or "").strip()
    probability_mode = config.get("probability_mode", PROBABILITY_MODE_CENTER)
    distance_threshold = float(config.get("distance_threshold") or 4.0)
    log_callback(f"[SRC] 概率计算方式: {_probability_mode_label(probability_mode)}")
    log_callback(f"[SRC] 矿点命中距离: {distance_threshold:g}")

    def _normalize_curve_points(src_curve, pac_curve):
        src_curve = list(src_curve or [])
        pac_curve = list(pac_curve or [])
        if not src_curve and not pac_curve:
            return [], []

        ordered = {}
        for item in src_curve:
            threshold = float(item.get("threshold", 0.0) or 0.0)
            ordered.setdefault(threshold, {})["src"] = float(item.get("sr", item.get("src", 0.0)) or 0.0)
        for item in pac_curve:
            threshold = float(item.get("threshold", 0.0) or 0.0)
            ordered.setdefault(threshold, {})["pac"] = float(item.get("paf", item.get("pac", 0.0)) or 0.0)

        thresholds = sorted(ordered.keys())
        pac_values = [max(0.0, min(1.0, float(ordered[t].get("pac", 0.0)))) for t in thresholds]
        src_values = [max(0.0, min(1.0, float(ordered[t].get("src", 0.0)))) for t in thresholds]
        return _prepare_src_curve(pac_values, src_values)

    def _prepare_src_curve(pac_values, src_values):
        pairs = []
        for pac, src in zip(pac_values or [], src_values or []):
            try:
                pairs.append(
                    (
                        max(0.0, min(1.0, float(pac))),
                        max(0.0, min(1.0, float(src))),
                    )
                )
            except (TypeError, ValueError):
                continue
        if not pairs:
            return [], []

        best_by_pac = {}
        for pac, src in pairs:
            best_by_pac[pac] = max(src, best_by_pac.get(pac, 0.0))

        if 0.0 not in best_by_pac:
            best_by_pac[0.0] = 0.0

        sorted_pairs = sorted(best_by_pac.items(), key=lambda item: item[0])
        return [pac for pac, _ in sorted_pairs], [src for _, src in sorted_pairs]

    def _hit_count_for_positions(selected_positions, deposit_coords, distance_threshold):
        selected_positions = np.asarray(selected_positions, dtype=np.float64)
        deposit_coords = np.asarray(deposit_coords, dtype=np.float64)
        if len(selected_positions) == 0 or len(deposit_coords) == 0:
            return 0

        try:
            from scipy.spatial import cKDTree  # type: ignore
        except Exception:
            cKDTree = None

        if cKDTree is not None and len(selected_positions) >= 10:
            tree = cKDTree(selected_positions[:, :2])
            hit_count = 0
            for deposit in deposit_coords[:, :2]:
                if tree.query_ball_point(deposit, r=float(distance_threshold)):
                    hit_count += 1
            return hit_count

        hit_count = 0
        for deposit in deposit_coords[:, :2]:
            distances = np.sqrt(np.sum((selected_positions[:, :2] - deposit) ** 2, axis=1))
            if len(distances) and float(np.min(distances)) <= float(distance_threshold):
                hit_count += 1
        return hit_count

    def _threshold_src_curve(positions, scores, deposit_coords, distance_threshold, threshold_step=0.001):
        positions = np.asarray(positions, dtype=np.float64)
        scores = np.asarray(scores, dtype=np.float64).reshape(-1)
        deposit_coords = np.asarray(deposit_coords, dtype=np.float64)
        if len(positions) == 0 or len(scores) == 0 or len(deposit_coords) == 0 or len(positions) != len(scores):
            return [], []

        finite_mask = np.isfinite(scores)
        positions = positions[finite_mask]
        scores = scores[finite_mask]
        if len(scores) == 0:
            return [], []

        total_area = len(scores)
        total_deposits = len(deposit_coords)
        pac_values = [0.0]
        src_values = [0.0]
        try:
            threshold_step = float(threshold_step or 0.001)
        except (TypeError, ValueError):
            threshold_step = 0.001
        if threshold_step <= 0 or threshold_step > 1:
            threshold_step = 0.001

        thresholds = np.arange(0.0, 1.0 + threshold_step / 2.0, threshold_step, dtype=np.float64)
        thresholds = np.unique(np.round(np.concatenate(([0.0, 1.0], thresholds)), 10))
        for threshold in thresholds:
            selected_mask = scores > float(threshold)
            selected_positions = positions[selected_mask]
            pac = float(np.sum(selected_mask) / total_area) if total_area else 0.0
            hit_count = _hit_count_for_positions(selected_positions, deposit_coords, distance_threshold)
            src = float(hit_count / total_deposits) if total_deposits else 0.0
            pac_values.append(pac)
            src_values.append(src)

        return _prepare_src_curve(pac_values, src_values)

    def _load_independent_curve_from_dat(model_path: str, model_name: str, deposit_path: str, area_path: str):
        if not SimpleDECGenerator:
            raise RuntimeError("未找到 DEC 模块，请确认 DEC.py 可用")
        temp_generator = SimpleDECGenerator(deposit_file=deposit_path)
        data = temp_generator.load_model_predictions(model_path, model_name)
        positions = np.column_stack(
            [
                np.asarray(data["x_coords"], dtype=np.float64) - 8.0,
                np.asarray(data["y_coords"], dtype=np.float64) - 8.0,
            ]
        )
        prediction_data = {
            "positions": positions,
            "confidences": np.asarray(data["probabilities"], dtype=np.float64).reshape(-1),
        }
        prediction_data = apply_probability_mode(prediction_data, probability_mode, log_callback, "SRC")
        calculator = GDRCalculator(
            deposit_file=deposit_path,
            distance_threshold=float(config.get("distance_threshold") or 4.0),
            confidence_threshold=0.5,
        )
        area_mask = calculator.load_test_area_mask(prediction_data, area_path)
        scoped_positions = positions[area_mask]
        scoped_confidences = prediction_data["confidences"][area_mask]
        if len(scoped_positions) == 0:
            raise ValueError(f"模型 {model_name} 的独立测试区内没有可用预测点。")
        deposit_coords = calculator.deposits_df[[calculator.x_col, calculator.y_col]].to_numpy(dtype=np.float64)
        pac_values, src_values = _threshold_src_curve(
            scoped_positions,
            scoped_confidences,
            deposit_coords,
            float(config.get("distance_threshold") or 4.0),
        )
        return {"name": model_name, "pac": pac_values, "src": src_values}

    def _load_curve_from_zone_statistics(zone_statistics_path: str, model_name: str):
        table = pd.read_csv(zone_statistics_path, encoding="utf-8-sig")
        if "test_sr" not in table.columns or "test_paf" not in table.columns:
            raise ValueError(f"zone_statistics.csv 缺少 test_sr/test_paf 列: {zone_statistics_path}")

        pac_values = [max(0.0, min(1.0, float(value))) for value in table["test_paf"].fillna(0.0).tolist()]
        src_values = [max(0.0, min(1.0, float(value))) for value in table["test_sr"].fillna(0.0).tolist()]
        pac_values, src_values = _prepare_src_curve(pac_values, src_values)
        return {
            "name": model_name,
            "pac": pac_values,
            "src": src_values,
        }

    def _load_curve_from_manifest(manifest_path: str, model_name: str | None = None):
        with open(manifest_path, "r", encoding="utf-8") as fh:
            manifest = json.load(fh)

        resolved_name = model_name or str(manifest.get("model_name") or Path(manifest_path).parent.name)
        zone_statistics_path = str(manifest.get("zone_statistics_path") or "").strip()
        if zone_statistics_path and os.path.exists(zone_statistics_path):
            return _load_curve_from_zone_statistics(zone_statistics_path, resolved_name)

        src_curve = manifest.get("src_curve") or []
        pac_curve = manifest.get("pac_curve") or []
        pac_values, src_values = _normalize_curve_points(src_curve, pac_curve)
        if pac_values or src_values:
            return {"name": resolved_name, "pac": pac_values, "src": src_values}

        prediction_header_path = str(manifest.get("prediction_header_path") or "").strip()
        if prediction_header_path and os.path.exists(prediction_header_path):
            with open(prediction_header_path, "r", encoding="utf-8") as fh:
                header = json.load(fh)
            pac_values, src_values = _normalize_curve_points(header.get("src_curve"), header.get("pac_curve"))
            if pac_values or src_values:
                return {"name": resolved_name, "pac": pac_values, "src": src_values}

        raise ValueError(f"清单中未找到可用的 SRC/PAC 曲线数据: {manifest_path}")

    def _load_curve_from_model_entry(model: dict):
        nonlocal generator
        source_type = str(model.get("source_type") or "").lower()
        model_name = str(model.get("plot_name") or model.get("name") or Path(str(model.get("path") or "")).stem)
        model_path = str(model.get("path") or "")

        if source_type == "zone_csv":
            if use_test_area:
                log_callback(f"[SRC] {model_name} 是已汇总曲线文件，未包含原始预测点，使用文件内已有曲线。")
            return _load_curve_from_zone_statistics(model_path, model_name)
        if source_type == "manifest":
            if use_test_area:
                log_callback(f"[SRC] {model_name} 是清单曲线来源，若清单未含原始预测点则使用已有曲线。")
            return _load_curve_from_manifest(model_path, model_name)
        if source_type == "trained_model":
            model_test_area_file = str(model.get("test_area_path") or test_area_file).strip()
            test_mineral_file = str(model.get("test_mineral_path") or deposit_file).strip()
            data_path = str(model.get("data_path") or "").strip()
            if not data_path:
                raise ValueError(f"模型 {model_name} 缺少完整H5路径，无法生成SRC曲线。")
            if use_test_area and not model_test_area_file:
                raise ValueError(f"模型 {model_name} 缺少测试区域文件，无法限定独立测试区。")
            if not test_mineral_file:
                raise ValueError(f"模型 {model_name} 缺少矿点文件，无法计算SRC。")

            log_callback(f"[SRC] 构建{'独立测试集' if use_test_area else '全区'}预测数据: {model_name}")
            prediction_data = build_prediction_data_from_model(
                {
                    "model_path": model_path,
                    "data_path": data_path,
                    "label_file": test_mineral_file,
                    "deposit_file": test_mineral_file,
                    "prediction_patch_stride": config.get("prediction_patch_stride"),
                },
                log_callback,
            )
            prediction_data = apply_probability_mode(prediction_data, probability_mode, log_callback, "SRC")

            distance_threshold = float(
                config.get("distance_threshold_override")
                or model.get("distance_threshold")
                or config.get("distance_threshold")
                or 4.0
            )
            calculator = GDRCalculator(
                deposit_file=test_mineral_file,
                distance_threshold=distance_threshold,
                confidence_threshold=0.5,
            )
            positions = calculator._positions_array(prediction_data)
            confidences = np.asarray(prediction_data["confidences"], dtype=np.float64).reshape(-1)
            if use_test_area:
                area_mask = calculator.load_test_area_mask(prediction_data, model_test_area_file)
                scoped_positions = positions[area_mask]
                scoped_confidences = confidences[area_mask]
                if len(scoped_positions) == 0:
                    raise ValueError(f"模型 {model_name} 的独立测试区内没有可用预测点。")
            else:
                scoped_positions = positions
                scoped_confidences = confidences

            deposit_coords = calculator.deposits_df[[calculator.x_col, calculator.y_col]].to_numpy(dtype=np.float64)
            pac_values, src_values = _threshold_src_curve(
                scoped_positions,
                scoped_confidences,
                deposit_coords,
                distance_threshold,
            )
            return {
                "name": model_name,
                "pac": pac_values,
            "src": src_values,
        }

        if use_test_area and source_type == "dat":
            if not test_area_file:
                raise ValueError("独立测试集SRC需要测试区域文件。")
            if not deposit_file:
                raise ValueError("独立测试集SRC需要测试矿点文件。")
            return _load_independent_curve_from_dat(model_path, model_name, deposit_file, test_area_file)

        if source_type == "prediction_h5":
            if not deposit_file:
                raise ValueError("H5预测结果需要矿点文件参与SRC计算。")
            if use_test_area and not test_area_file:
                raise ValueError("独立测试集SRC需要测试区域文件。")
            calculator = GDRCalculator(
                deposit_file=deposit_file,
                distance_threshold=float(config.get("distance_threshold") or 4.0),
                confidence_threshold=0.5,
            )
            prediction_data = calculator.load_predictions(model_path)
            prediction_data = apply_probability_mode(prediction_data, probability_mode, log_callback, "SRC")
            positions = calculator._positions_array(prediction_data)
            confidences = np.asarray(prediction_data["confidences"], dtype=np.float64).reshape(-1)
            if use_test_area:
                area_mask = calculator.load_test_area_mask(prediction_data, test_area_file)
                scoped_positions = positions[area_mask]
                scoped_confidences = confidences[area_mask]
                if len(scoped_positions) == 0:
                    raise ValueError(f"预测H5 {model_name} 的独立测试区内没有可用预测点。")
            else:
                scoped_positions = positions
                scoped_confidences = confidences
            deposit_coords = calculator.deposits_df[[calculator.x_col, calculator.y_col]].to_numpy(dtype=np.float64)
            pac_values, src_values = _threshold_src_curve(
                scoped_positions,
                scoped_confidences,
                deposit_coords,
                float(config.get("distance_threshold") or 4.0),
            )
            return {
                "name": model_name,
                "pac": pac_values,
                "src": src_values,
            }

        if generator is None:
            if not SimpleDECGenerator:
                raise RuntimeError("未找到 DEC 模块，请确认 DEC.py 可用")
            generator = SimpleDECGenerator(deposit_file=deposit_file)
        data = generator.load_model_predictions(model_path, model_name)
        prediction_data = {
            "positions": np.column_stack(
                [
                    np.asarray(data["x_coords"], dtype=np.float64) - 8.0,
                    np.asarray(data["y_coords"], dtype=np.float64) - 8.0,
                ]
            ),
            "confidences": np.asarray(data["probabilities"], dtype=np.float64).reshape(-1),
        }
        prediction_data = apply_probability_mode(prediction_data, probability_mode, log_callback, "SRC")
        deposit_coords = generator.deposits_df[[generator.x_col, generator.y_col]].to_numpy(dtype=np.float64)
        pac_values, src_values = _threshold_src_curve(
            prediction_data["positions"],
            prediction_data["confidences"],
            deposit_coords,
            float(config.get("distance_threshold") or 4.0),
        )
        return {
            "name": model_name,
            "pac": pac_values,
            "src": src_values,
        }

    generator = None
    curves = []
    color_palette = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                    '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']

    for model in models:
        source_type = str(model.get("source_type") or "dat").lower()
        log_prefix = "[SRC]"
        if source_type == "dat":
            log_callback(f"{log_prefix} 从 DAT 计算曲线: {model['name']}")
        else:
            log_callback(f"{log_prefix} 加载新模式曲线: {model['name']}")
        curve_data = _load_curve_from_model_entry(model)
        curves.append({
            "name": curve_data["name"],
            "group_name": str(model.get("curve_group_name") or curve_data["name"]),
            "path": str(model.get("path") or ""),
            "source_type": str(model.get("source_type") or ""),
            "color": color_palette[len(curves) % len(color_palette)],
            "pac": curve_data["pac"],
            "src": curve_data["src"],
        })

    return {"curves": curves}


def _threshold_values(step: float) -> np.ndarray:
    step = float(step or 0.01)
    if step <= 0 or step > 1:
        raise ValueError("阈值步长需要在 0-1 之间")
    values = np.arange(0.0, 1.0 + step * 0.5, step, dtype=np.float64)
    values = np.unique(np.clip(values, 0.0, 1.0))
    if values[-1] < 1.0:
        values = np.append(values, 1.0)
    return values


def _pa_intersection(curve: list[dict]) -> dict:
    if not curve:
        return {}
    thresholds = np.asarray([item["threshold"] for item in curve], dtype=np.float64)
    sr_values = np.asarray([item["sr"] for item in curve], dtype=np.float64)
    paf_values = np.asarray([item["paf"] for item in curve], dtype=np.float64)
    diff = sr_values + paf_values - 1.0
    nearest_i = int(np.nanargmin(np.abs(diff)))
    nearest = {
        "nearest_threshold": float(thresholds[nearest_i]),
        "nearest_sr": float(sr_values[nearest_i]),
        "nearest_paf": float(paf_values[nearest_i]),
        "nearest_gap": float(diff[nearest_i]),
    }
    eps = 1e-12
    for i in range(len(diff) - 1):
        d0, d1 = diff[i], diff[i + 1]
        if not np.isfinite(d0) or not np.isfinite(d1):
            continue
        if abs(d0) < eps:
            return {
                "type": "exact",
                "threshold": float(thresholds[i]),
                "sr": float(sr_values[i]),
                "paf": float(paf_values[i]),
                "pa_intersection_y": float(sr_values[i]),
                **nearest,
            }
        if d0 * d1 < 0:
            weight = -d0 / (d1 - d0)
            sr = sr_values[i] + weight * (sr_values[i + 1] - sr_values[i])
            paf = paf_values[i] + weight * (paf_values[i + 1] - paf_values[i])
            return {
                "type": "linear_interpolation",
                "threshold": float(thresholds[i] + weight * (thresholds[i + 1] - thresholds[i])),
                "sr": float(sr),
                "paf": float(paf),
                "pa_intersection_y": float(sr),
                **nearest,
            }
    return {
        "type": "nearest_only_no_crossing",
        "threshold": nearest["nearest_threshold"],
        "sr": nearest["nearest_sr"],
        "paf": nearest["nearest_paf"],
        "pa_intersection_y": nearest["nearest_sr"],
        **nearest,
    }


def _build_pa_curve_from_prediction_data(
    prediction_data: dict,
    deposit_file: str,
    distance_threshold: float,
    threshold_step: float,
    use_test_area: bool = False,
    test_area_file: str = "",
) -> tuple[list[dict], dict]:
    from GDR_calculator import GDRCalculator

    calculator = GDRCalculator(
        deposit_file=deposit_file,
        distance_threshold=float(distance_threshold),
        confidence_threshold=0.5,
    )
    positions = calculator._positions_array(prediction_data)
    confidences = np.asarray(prediction_data["confidences"], dtype=np.float64).reshape(-1)
    if len(positions) != len(confidences):
        raise ValueError(f"预测坐标与置信度数量不一致: positions={len(positions)}, confidences={len(confidences)}")

    if use_test_area:
        area_mask = calculator.load_test_area_mask(prediction_data, test_area_file)
        area_mask = np.asarray(area_mask, dtype=bool).reshape(-1)
        if len(area_mask) != len(confidences):
            raise ValueError(f"测试区域mask长度不匹配: mask={len(area_mask)}, predictions={len(confidences)}")
        scoped_positions = positions[area_mask]
        scoped_confidences = confidences[area_mask]
    else:
        scoped_positions = positions
        scoped_confidences = confidences

    if len(scoped_positions) == 0:
        raise ValueError("P-A 曲线计算范围内没有可用预测点。")

    deposit_coords = calculator.deposits_df[[calculator.x_col, calculator.y_col]].to_numpy(dtype=np.float64)
    total_deposits = len(deposit_coords)
    if total_deposits == 0:
        raise ValueError("矿点文件中没有可用矿点坐标。")

    curve = []
    for threshold in _threshold_values(threshold_step):
        selected = scoped_positions[scoped_confidences > threshold]
        paf = len(selected) / len(scoped_positions)
        hit_count, _, _ = calculator._deposit_hit_stats(selected, deposit_coords)
        sr = hit_count / total_deposits if total_deposits else 0.0
        curve.append(
            {
                "threshold": float(threshold),
                "sr": float(sr),
                "paf": float(paf),
                "ei": float(sr / paf) if paf > 0 else np.nan,
                "hit_deposits": int(hit_count),
                "total_deposits": int(total_deposits),
                "high_potential_count": int(len(selected)),
                "area_count": int(len(scoped_positions)),
            }
        )
    return curve, _pa_intersection(curve)


def _load_pa_curve_from_zone_statistics(zone_statistics_path: str, model_name: str) -> dict:
    table = pd.read_csv(zone_statistics_path, encoding="utf-8-sig")
    threshold_col = next((col for col in ("threshold", "probability_threshold", "阈值") if col in table.columns), None)
    sr_col = next((col for col in ("test_sr", "SR", "sr", "gdr") if col in table.columns), None)
    paf_col = next((col for col in ("test_paf", "PAF", "paf", "pa", "par") if col in table.columns), None)
    if threshold_col is None or sr_col is None or paf_col is None:
        raise ValueError(f"zone_statistics.csv 缺少 threshold/SR/PAF 列: {zone_statistics_path}")
    curve = []
    for _, row in table.iterrows():
        try:
            threshold = float(row[threshold_col])
            sr = max(0.0, min(1.0, float(row[sr_col])))
            paf = max(0.0, min(1.0, float(row[paf_col])))
        except (TypeError, ValueError):
            continue
        curve.append({"threshold": threshold, "sr": sr, "paf": paf, "ei": sr / paf if paf > 0 else np.nan})
    curve = sorted(curve, key=lambda item: item["threshold"])
    if not curve:
        raise ValueError(f"zone_statistics.csv 中没有可用 P-A 曲线数据: {zone_statistics_path}")
    return {"name": model_name, "curve": curve, "intersection": _pa_intersection(curve)}


def _load_pa_curve_from_manifest(manifest_path: str, model_name: str | None = None) -> dict:
    with open(manifest_path, "r", encoding="utf-8") as fh:
        manifest = json.load(fh)
    resolved_name = model_name or str(manifest.get("model_name") or Path(manifest_path).parent.name)
    zone_statistics_path = str(manifest.get("zone_statistics_path") or "").strip()
    if zone_statistics_path and not os.path.isabs(zone_statistics_path):
        zone_statistics_path = os.path.join(os.path.dirname(manifest_path), zone_statistics_path)
    if zone_statistics_path and os.path.exists(zone_statistics_path):
        return _load_pa_curve_from_zone_statistics(zone_statistics_path, resolved_name)
    raise ValueError(f"清单中未找到可用于 P-A 曲线的 zone_statistics.csv: {manifest_path}")


def run_pa_task(models: list[dict], log_callback, deposit_file: str, config: dict | None = None):
    from GDR_calculator import GDRCalculator

    config = dict(config or {})
    use_test_area = bool(config.get("use_test_area"))
    test_area_file = str(config.get("test_area_file") or "").strip()
    probability_mode = config.get("probability_mode", PROBABILITY_MODE_CENTER)
    threshold_step = float(config.get("threshold_step") or 0.01)
    distance_threshold = float(config.get("distance_threshold") or 4.0)
    log_callback(f"[P-A] 概率计算方式: {_probability_mode_label(probability_mode)}")

    curves = []
    color_palette = ['#d62728', '#2ca02c', '#1f77b4', '#ff7f0e', '#9467bd',
                     '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']

    for model in models:
        source_type = str(model.get("source_type") or "dat").lower()
        model_name = str(model.get("plot_name") or model.get("name") or Path(str(model.get("path") or "")).stem)
        model_path = str(model.get("path") or "")
        log_callback(f"[P-A] 生成曲线: {model_name}")

        if source_type == "zone_csv":
            curve_data = _load_pa_curve_from_zone_statistics(model_path, model_name)
        elif source_type == "manifest":
            curve_data = _load_pa_curve_from_manifest(model_path, model_name)
        elif source_type == "trained_model":
            model_test_area_file = str(model.get("test_area_path") or test_area_file).strip()
            test_mineral_file = str(model.get("test_mineral_path") or deposit_file).strip()
            data_path = str(model.get("data_path") or "").strip()
            if not data_path:
                raise ValueError(f"模型 {model_name} 缺少完整H5路径，无法生成P-A曲线。")
            if use_test_area and not model_test_area_file:
                raise ValueError(f"模型 {model_name} 缺少测试区域文件，无法限定独立测试区。")
            if not test_mineral_file:
                raise ValueError(f"模型 {model_name} 缺少矿点文件，无法计算P-A曲线。")
            prediction_data = build_prediction_data_from_model(
                {
                    "model_path": model_path,
                    "data_path": data_path,
                    "label_file": test_mineral_file,
                    "deposit_file": test_mineral_file,
                    "prediction_patch_stride": config.get("prediction_patch_stride"),
                },
                log_callback,
            )
            prediction_data = apply_probability_mode(prediction_data, probability_mode, log_callback, "P-A")
            curve, intersection = _build_pa_curve_from_prediction_data(
                prediction_data,
                test_mineral_file,
                distance_threshold,
                threshold_step,
                use_test_area=use_test_area,
                test_area_file=model_test_area_file,
            )
            curve_data = {"name": model_name, "curve": curve, "intersection": intersection}
        else:
            if not deposit_file:
                raise ValueError("DAT/H5预测结果需要矿点文件参与P-A计算。")
            calculator = GDRCalculator(
                deposit_file=deposit_file,
                distance_threshold=distance_threshold,
                confidence_threshold=0.5,
            )
            prediction_data = calculator.load_predictions(model_path)
            prediction_data = apply_probability_mode(prediction_data, probability_mode, log_callback, "P-A")
            curve, intersection = _build_pa_curve_from_prediction_data(
                prediction_data,
                deposit_file,
                distance_threshold,
                threshold_step,
                use_test_area=use_test_area,
                test_area_file=test_area_file,
            )
            curve_data = {"name": model_name, "curve": curve, "intersection": intersection}

        curve_data["color"] = color_palette[len(curves) % len(color_palette)]
        curves.append(curve_data)

    return {"curves": curves}


def run_shap_task(config: dict, log_callback):
    from evaluate_shap_cnnt import explain_cnnt_with_shap
    shap_log_lines = []

    class _ShapLogWriter:
        def __init__(self, prefix):
            self.prefix = prefix
            self._buffer = ""

        def write(self, message):
            if not message:
                return
            self._buffer += message
            while "\n" in self._buffer:
                line, self._buffer = self._buffer.split("\n", 1)
                line = line.strip()
                if line:
                    formatted = f"{self.prefix}{line}"
                    shap_log_lines.append(formatted)
                    log_callback(formatted)

        def flush(self):
            content = self._buffer.strip()
            if content:
                formatted = f"{self.prefix}{content}"
                shap_log_lines.append(formatted)
                log_callback(formatted)
            self._buffer = ""

    stdout_writer = _ShapLogWriter("[SHAP] ")
    stderr_writer = _ShapLogWriter("[SHAP-ERR] ")

    def _make_args(model_path, output_dir):
        return SimpleNamespace(
            model_path=model_path,
            model_type=config["model_type"],
            input=config["input_data"],
            label_file=config["label_file"],
            test_area_file=config.get("test_area_file") or None,
            use_test_area=bool(config.get("use_test_area")),
            use_prediction_dataset=bool(config.get("use_prediction_dataset")),
            norm_params_path=config.get("norm_params_path") or None,
            patch_size=config.get("patch_size"),
            patch_stride=config.get("patch_stride"),
            reflect_padding=bool(config.get("reflect_padding")),
            selected_channels=config.get("selected_channels") or "",
            prior=config["prior"],
            num_background_samples=config["num_background_samples"],
            num_test_samples=config["num_test_samples"],
            device=config["device"],
            figure_dpi=config["figure_dpi"],
            channel_importance_figsize=config["channel_importance_figsize"],
            summary_plot_figsize=config["summary_plot_figsize"],
            kernel_nsamples=100,
            output_dir=output_dir,
            custom_channel_names=config.get("custom_channel_names", ""),
            importance_xlabel=config.get("importance_xlabel", "Mean |SHAP|"),
            summary_xlabel=config.get("summary_xlabel", "SHAP value (impact on model output)"),
        )

    model_paths = list(config.get("model_paths") or [config["model_path"]])
    base_output_dir = config.get("output_dir") or None
    if base_output_dir:
        base_output_dir = os.path.abspath(base_output_dir)

    def _run_one(model_path, output_dir):
        args = _make_args(model_path, output_dir)
        with contextlib.redirect_stdout(stdout_writer), contextlib.redirect_stderr(stderr_writer):
            explain_cnnt_with_shap(args)
        stdout_writer.flush()
        stderr_writer.flush()
        resolved_output_dir = (
            os.path.abspath(args.output_dir)
            if args.output_dir
            else os.path.join(os.path.dirname(model_path), "shap_explanations")
        )
        result = {
            "model_path": model_path,
            "output_dir": resolved_output_dir,
            "channel_plot": os.path.join(resolved_output_dir, "shap_channel_importance.png"),
            "summary_plot": os.path.join(resolved_output_dir, "shap_summary_plot.png"),
            "importance_csv": os.path.join(resolved_output_dir, "shap_channel_importance.csv"),
            "summary_values_csv": os.path.join(resolved_output_dir, "shap_summary_values.csv"),
        }
        missing = [
            path for path in (result["channel_plot"], result["summary_plot"], result["importance_csv"])
            if not os.path.exists(path) or os.path.getsize(path) == 0
        ]
        if missing:
            recent_logs = "\n".join(shap_log_lines[-15:]) if shap_log_lines else "无 SHAP 日志输出"
            raise RuntimeError(
                "SHAP任务结束但未生成完整输出文件。\n"
                f"缺失文件: {', '.join(os.path.basename(path) for path in missing)}\n"
                "请根据以下最近日志定位原因:\n"
                f"{recent_logs}"
            )
        return result

    if len(model_paths) == 1:
        output_dir = base_output_dir
        return _run_one(model_paths[0], output_dir)

    repeat_output_dir = base_output_dir or os.path.join(os.path.dirname(model_paths[0]), "shap_repeat_summary")
    repeat_output_dir = os.path.abspath(repeat_output_dir)
    os.makedirs(repeat_output_dir, exist_ok=True)
    log_callback(f"[SHAP] 多模型SHAP汇总: {len(model_paths)} 个模型")
    run_results = []
    importance_tables = []
    summary_value_tables = []
    for index, model_path in enumerate(model_paths, start=1):
        run_output_dir = os.path.join(repeat_output_dir, f"model_{index:03d}_{Path(model_path).parent.name}")
        log_callback(f"[SHAP] 运行第 {index}/{len(model_paths)} 个模型: {model_path}")
        result = _run_one(model_path, run_output_dir)
        run_results.append(result)
        table = pd.read_csv(result["importance_csv"])
        table["channel"] = table["channel"].astype(str)
        table["model_index"] = index
        table["model_path"] = model_path
        importance_tables.append(table)
        summary_values_csv = result.get("summary_values_csv")
        if summary_values_csv and os.path.exists(summary_values_csv):
            values_table = pd.read_csv(summary_values_csv)
            values_table["channel"] = values_table["channel"].astype(str)
            values_table["model_index"] = index
            values_table["model_path"] = model_path
            summary_value_tables.append(values_table)

    all_importance = pd.concat(importance_tables, ignore_index=True)
    all_csv = os.path.join(repeat_output_dir, "shap_channel_importance_all_runs.csv")
    all_importance.to_csv(all_csv, index=False, encoding="utf-8-sig")
    summary = (
        all_importance.groupby("channel", as_index=False)["importance"]
        .agg(["mean", "std", "median", "count"])
        .reset_index()
        .sort_values("mean", ascending=False)
    )
    summary_csv = os.path.join(repeat_output_dir, "shap_channel_importance_summary.csv")
    summary.to_csv(summary_csv, index=False, encoding="utf-8-sig")

    fig_height = max(4.0, min(12.0, 0.35 * len(summary) + 1.5))
    plt.figure(figsize=(10, fig_height))
    y = np.arange(len(summary))
    show_error_bars = bool(config.get("show_error_bars", True))
    importance_xlabel = str(config.get("importance_xlabel") or "Mean |SHAP|")
    summary_xlabel = str(config.get("summary_xlabel") or "SHAP value (impact on model output)")
    xerr = summary["std"].fillna(0.0) if show_error_bars else None
    plt.barh(y, summary["mean"], xerr=xerr, alpha=0.75, color="#4e79a7")
    plt.yticks(y, summary["channel"])
    plt.gca().invert_yaxis()
    plt.xlabel(importance_xlabel)
    title_suffix = "with std" if show_error_bars else "no std bars"
    plt.title(f"Mean SHAP Channel Importance ({len(model_paths)} models, {title_suffix})")
    plt.tight_layout()
    summary_plot_name = (
        "shap_channel_importance_mean_errorbar.png"
        if show_error_bars
        else "shap_channel_importance_mean_no_errorbar.png"
    )
    summary_plot = os.path.join(repeat_output_dir, summary_plot_name)
    plt.savefig(summary_plot, dpi=int(config.get("figure_dpi") or 300), bbox_inches="tight")
    plt.close()

    summary_channel_order = [str(value) for value in summary["channel"].tolist()]
    aggregate_summary_plot = os.path.join(repeat_output_dir, "shap_summary_plot_mean_order.png")
    all_summary_values_csv = ""
    if summary_value_tables:
        try:
            import shap

            all_summary_values = pd.concat(summary_value_tables, ignore_index=True)
            all_summary_values_csv = os.path.join(repeat_output_dir, "shap_summary_values_all_runs.csv")
            all_summary_values.to_csv(all_summary_values_csv, index=False, encoding="utf-8-sig")

            pivot_index = ["model_index", "sample_order"]
            shap_wide = all_summary_values.pivot_table(
                index=pivot_index,
                columns="channel",
                values="shap_value",
                aggfunc="mean",
            )
            feature_wide = all_summary_values.pivot_table(
                index=pivot_index,
                columns="channel",
                values="feature_value",
                aggfunc="mean",
            )
            common_index = shap_wide.index.intersection(feature_wide.index)
            present_order = [
                channel
                for channel in summary_channel_order
                if channel in shap_wide.columns and channel in feature_wide.columns
            ]
            if not present_order:
                raise ValueError("汇总SHAP明细中没有与通道重要性表匹配的通道名称")

            shap_values_2d = shap_wide.loc[common_index, present_order].to_numpy(dtype=np.float64)
            feature_values_2d = feature_wide.loc[common_index, present_order].to_numpy(dtype=np.float64)
            valid_rows = np.isfinite(shap_values_2d).all(axis=1) & np.isfinite(feature_values_2d).all(axis=1)
            shap_values_2d = shap_values_2d[valid_rows]
            feature_values_2d = feature_values_2d[valid_rows]
            if shap_values_2d.size == 0:
                raise ValueError("汇总SHAP明细中没有可绘制的有限数值")

            shap.summary_plot(
                shap_values_2d,
                feature_values_2d,
                feature_names=present_order,
                show=False,
                plot_type="dot",
                color_bar=True,
                max_display=len(present_order),
                sort=False,
            )
            ax = plt.gca()
            for spine in ax.spines.values():
                spine.set_visible(True)
                spine.set_linewidth(1.0)
                spine.set_color("black")
            ax.set_xlabel(summary_xlabel)
            plt.gcf().set_size_inches(*config.get("summary_plot_figsize", (10, 8)))
            plt.savefig(
                aggregate_summary_plot,
                dpi=int(config.get("figure_dpi") or 300),
                bbox_inches="tight",
                facecolor="white",
            )
            plt.close()
            log_callback(f"[SHAP] 多模型summary图已按汇总均值排序: {aggregate_summary_plot}")
        except Exception as exc:
            plt.close()
            aggregate_summary_plot = run_results[0]["summary_plot"]
            log_callback(f"[SHAP] 生成多模型summary汇总图失败，回退到首个模型summary图: {exc}")
    else:
        aggregate_summary_plot = run_results[0]["summary_plot"]
        log_callback("[SHAP] 未找到summary明细，回退到首个模型summary图。")

    representative = run_results[0]
    log_callback(f"[SHAP] 多模型SHAP汇总完成: {summary_csv}")
    return {
        "output_dir": repeat_output_dir,
        "channel_plot": summary_plot,
        "summary_plot": aggregate_summary_plot,
        "importance_csv": summary_csv,
        "all_importance_csv": all_csv,
        "all_summary_values_csv": all_summary_values_csv,
        "run_results": run_results,
        "is_repeat_summary": True,
    }


class EvaluationWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PU Learning 模型评估工具")
        self.resize(1300, 860)

        self.log_queue = queue.Queue()
        self.active_thread = None
        self.dec_available_models = []
        self.pa_available_models = []
        self.gdr_folder_model_entries = []
        self.gdr_available_model_name_counts = {}
        self.dec_last_result = None
        self.pa_last_result = None
        self.shap_plot_paths = {"channel": None, "summary": None}
        self.shap_pixmaps = {"channel": None, "summary": None}
        self.confusion_matrix_data = None
        self.shap_last_result = None

        self._init_defaults()
        self._build_ui()

        self.log_timer = QTimer(self)
        self.log_timer.timeout.connect(self._process_log_queue)
        self.log_timer.start(120)

    def _init_defaults(self):
        default_deposit = os.path.join(BASE_DIR, "deposit.xlsx")
        default_model_dir = os.path.join(BASE_DIR, "result", "cnnt-9-0.2-10%")

        self.prediction_path = ""
        self.eval_model_path = ""
        self.gdr_model_root_path = ""
        self.eval_data_path = ""
        self.gdr_deposit_path = default_deposit
        self.distance_threshold = "4"
        self.confidence_threshold = "0.5"
        self.prior_value = "0.2"
        self.gdr_prediction_stride = ""
        self.gdr_probability_mode = PROBABILITY_MODE_CENTER
        self.test_area_path = ""
        self.threshold_step = "0.01"
        self.src_h5_path = ""
        self.src_test_area_path = ""
        self.src_prediction_stride = ""
        self.src_distance_threshold = "4"
        self.src_probability_mode = PROBABILITY_MODE_CENTER
        self.src_show_uncertainty_band = False
        self.src_uncertainty_mode = "std"
        self.dec_deposit_path = default_deposit
        self.pa_h5_path = ""
        self.pa_test_area_path = ""
        self.pa_prediction_stride = ""
        self.pa_probability_mode = PROBABILITY_MODE_CENTER
        self.pa_deposit_path = default_deposit
        self.pa_threshold_step = "0.01"
        self.pa_distance_threshold = "4"
        self.shap_model_path = os.path.join(default_model_dir, "model.pth")
        self.shap_model_type = "cnnt"
        self.shap_input_path = os.path.join(BASE_DIR, "data", "windows_combined_data.h5")
        self.shap_label_path = os.path.join(BASE_DIR, "data", "deposit_labels.h5")
        self.shap_test_area_path = ""
        self.shap_background = "50"
        self.shap_test = "200"
        self.shap_prior = "0.2"
        self.shap_prediction_stride = ""
        self.shap_show_error_bars = True
        self.shap_importance_xlabel = "Mean |SHAP|"
        self.shap_summary_xlabel = "SHAP value (impact on model output)"
        self.shap_channel_name_overrides = {}
        self.shap_device = "cuda"
        self.shap_output_dir = os.path.join(BASE_DIR, "shap_outputs")
        self.shap_output_text = "尚未运行 SHAP 分析"

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(12, 12, 12, 12)
        root_layout.setSpacing(10)

        self.tabs = QTabWidget()
        root_layout.addWidget(self.tabs, 1)

        self._build_gdr_tab()
        self._build_dec_tab()
        self._build_pa_tab()
        self._build_shap_tab()

        status_row = QHBoxLayout()
        status_row.addStretch(1)
        self.status_label = QLabel("就绪")
        status_row.addWidget(self.status_label)
        root_layout.addLayout(status_row)

    def _build_gdr_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(10)
        layout.addWidget(QLabel("可选择模型文件现场预测评估；也可载入已有预测结果文件（与模型文件二选一）。"))

        param_group = QGroupBox("参数设置")
        form = QFormLayout(param_group)
        form.setLabelAlignment(Qt.AlignRight)

        self.prediction_edit = self._create_path_input(file_filter="预测文件 (*.h5 *.hdf5 *.dat);;所有文件 (*.*)", default=self.prediction_path)
        form.addRow("已有预测结果文件（可选）", self.prediction_edit['container'])

        self.eval_model_edit = self._create_path_input(file_filter="模型文件 (*.pth *.pkl);;所有文件 (*.*)", default=self.eval_model_path, multi_files=True)
        form.addRow("模型文件（可选）", self.eval_model_edit['container'])

        self.eval_data_edit = self._create_path_input(file_filter="完整特征H5 (*.h5 *.hdf5);;所有文件 (*.*)", default=self.eval_data_path)
        form.addRow("完整特征H5", self.eval_data_edit['container'])

        self.gdr_deposit_edit = self._create_path_input(file_filter="矿点文件 (*.txt *.dat *.csv *.tsv *.xlsx *.xls);;所有文件 (*.*)", default=self.gdr_deposit_path)
        form.addRow("矿点文件", self.gdr_deposit_edit['container'])

        self.distance_edit = self._create_line_edit(self.distance_threshold)
        form.addRow("距离阈值", self.distance_edit)

        self.confidence_edit = self._create_line_edit(self.confidence_threshold)
        form.addRow("置信度阈值", self.confidence_edit)

        self.prior_edit = self._create_line_edit(self.prior_value)
        form.addRow("先验概率", self.prior_edit)

        self.gdr_prediction_stride_edit = self._create_line_edit(self.gdr_prediction_stride)
        self.gdr_prediction_stride_edit.setPlaceholderText("留空则沿用训练步长，可填 1 提高评估分辨率")
        self.gdr_prediction_stride_edit.setToolTip("从模型文件现场预测评估时使用；窗口大小仍沿用训练参数。")
        form.addRow("评估步长", self.gdr_prediction_stride_edit)

        self.gdr_probability_mode_combo = QComboBox()
        for label, value in PROBABILITY_MODE_CHOICES.items():
            self.gdr_probability_mode_combo.addItem(label, value)
        self.gdr_probability_mode_combo.setCurrentIndex(
            self.gdr_probability_mode_combo.findData(self.gdr_probability_mode)
        )
        form.addRow("概率计算方式", self.gdr_probability_mode_combo)

        self.independent_test_check = QCheckBox("独立测试集评估（只用测试矿点和测试区域）")
        self.independent_test_check.setChecked(False)
        form.addRow("评估模式", self.independent_test_check)

        self.threshold_strategy_combo = QComboBox()
        self.threshold_strategy_combo.addItems(["EI 最大", "手动阈值"])
        self.threshold_strategy_combo.setCurrentText("EI 最大")
        form.addRow("独立测试阈值策略", self.threshold_strategy_combo)

        self.test_area_edit = self._create_path_input(
            file_filter="测试区域文件 (*.h5 *.hdf5 *.npy *.npz *.txt *.dat *.csv *.tsv *.xlsx *.xls);;所有文件 (*.*)",
            default=self.test_area_path,
        )
        form.addRow("测试区域文件", self.test_area_edit['container'])

        self.threshold_step_edit = self._create_line_edit(self.threshold_step)
        form.addRow("阈值扫描步长", self.threshold_step_edit)

        layout.addWidget(param_group)

        folder_group = QGroupBox("文件夹批量评估")
        folder_form = QFormLayout(folder_group)
        folder_form.setLabelAlignment(Qt.AlignRight)

        self.gdr_model_root_edit = self._create_path_input(directory=True, default=self.gdr_model_root_path)
        folder_form.addRow("模型根目录", self.gdr_model_root_edit["container"])

        self.gdr_model_name_combo = QComboBox()
        self.gdr_model_name_combo.setEditable(True)
        self.gdr_model_name_combo.addItems([
            FIVE_FOLD_MODEL_LABEL,
            PKL_FIVE_FOLD_MODEL_LABEL,
            *FOLD_MODEL_NAMES,
            *PKL_FOLD_MODEL_NAMES,
            "best_model.pth",
            "best_model.pkl",
            "model.pth",
            "model.pkl",
        ])
        model_name_row = QWidget()
        model_name_layout = QHBoxLayout(model_name_row)
        model_name_layout.setContentsMargins(0, 0, 0, 0)
        model_name_layout.addWidget(self.gdr_model_name_combo, 1)
        select_names_btn = QPushButton("选择文件名...")
        select_names_btn.clicked.connect(self._select_gdr_model_names)
        model_name_layout.addWidget(select_names_btn)
        folder_form.addRow("模型文件名", model_name_row)

        folder_btn_row = QWidget()
        folder_btn_layout = QHBoxLayout(folder_btn_row)
        folder_btn_layout.setContentsMargins(0, 0, 0, 0)
        scan_folder_btn = QPushButton("扫描子文件夹")
        scan_folder_btn.clicked.connect(self._scan_gdr_model_folders)
        folder_btn_layout.addWidget(scan_folder_btn)
        apply_folder_btn = QPushButton("使用所选模型")
        apply_folder_btn.clicked.connect(self._apply_gdr_folder_models)
        folder_btn_layout.addWidget(apply_folder_btn)
        run_folder_btn = QPushButton("批量评估并保存结果")
        run_folder_btn.clicked.connect(self._run_gdr_folder_batch_csv)
        folder_btn_layout.addWidget(run_folder_btn)
        folder_btn_layout.addStretch(1)
        self.gdr_folder_status_label = QLabel("尚未扫描")
        self.gdr_folder_status_label.setStyleSheet("color: #666666;")
        folder_btn_layout.addWidget(self.gdr_folder_status_label)
        folder_form.addRow("批量操作", folder_btn_row)

        layout.addWidget(folder_group)

        run_btn = QPushButton("模型评估")
        run_btn.clicked.connect(self._run_gdr)
        layout.addWidget(run_btn)

        results_container = QHBoxLayout()
        layout.addLayout(results_container, 1)

        left_panel = QVBoxLayout()
        results_container.addLayout(left_panel, 1)

        metrics_group = QGroupBox("评估指标")
        metrics_layout = QVBoxLayout(metrics_group)
        self.metrics_table = QTableWidget(0, 2)
        self.metrics_table.setHorizontalHeaderLabels(["指标", "数值"])
        self.metrics_table.horizontalHeader().setStretchLastSection(True)
        self.metrics_table.verticalHeader().setVisible(False)
        self.metrics_table.setEditTriggers(QTableWidget.NoEditTriggers)
        metrics_layout.addWidget(self.metrics_table)
        left_panel.addWidget(metrics_group)

        log_group = QGroupBox("运行日志")
        log_layout = QVBoxLayout(log_group)
        toolbar = QHBoxLayout()
        toolbar.addStretch(1)
        clear_btn = QPushButton("清空日志")
        clear_btn.clicked.connect(self._clear_log)
        toolbar.addWidget(clear_btn)
        log_layout.addLayout(toolbar)
        self.log_view = QPlainTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setMaximumBlockCount(5000)
        log_layout.addWidget(self.log_view)
        left_panel.addWidget(log_group, 1)

        right_panel = QVBoxLayout()
        results_container.addLayout(right_panel, 1)

        confusion_group = QGroupBox("混淆矩阵")
        confusion_layout = QVBoxLayout(confusion_group)
        self.confusion_fig = Figure(figsize=(4, 4))
        self.confusion_ax = self.confusion_fig.add_subplot(111)
        self.confusion_canvas = FigureCanvasQTAgg(self.confusion_fig)
        confusion_layout.addWidget(self.confusion_canvas)
        view_confusion_btn = QPushButton("查看大图")
        view_confusion_btn.clicked.connect(self._show_confusion_fullscreen)
        confusion_layout.addWidget(view_confusion_btn)
        self.confusion_colorbar = None
        right_panel.addWidget(confusion_group, 1)

        self._update_metrics_display(None)
        self._update_confusion_matrix(None)

        self.tabs.addTab(tab, "模型评估")

    def _build_dec_tab(self):
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.addWidget(QLabel("通过“添加文件...”按钮选择训练模型(.pth)、DAT、rebuild_manifest.json、model_packages.json 或 zone_statistics.csv，生成 SRC/PAR 曲线。"))

        dec_group = QGroupBox("SRC参数")
        form = QFormLayout(dec_group)
        form.setLabelAlignment(Qt.AlignRight)
        self.src_h5_edit = self._create_path_input(file_filter="完整H5 (*.h5 *.hdf5);;所有文件 (*.*)", default=self.src_h5_path)
        form.addRow("完整H5(可选)", self.src_h5_edit['container'])
        self.src_prediction_stride_edit = self._create_line_edit(self.src_prediction_stride)
        self.src_prediction_stride_edit.setPlaceholderText("留空则沿用训练步长，可填 1 提高SRC分辨率")
        self.src_prediction_stride_edit.setToolTip("从训练模型临时生成SRC预测数据时使用；窗口大小仍沿用训练参数。")
        form.addRow("SRC预测步长", self.src_prediction_stride_edit)
        self.src_distance_edit = self._create_line_edit(self.src_distance_threshold)
        self.src_distance_edit.setPlaceholderText("默认 4，单位为坐标系长度")
        self.src_distance_edit.setToolTip("矿点到入选高潜力预测单元的距离不超过该值时计为命中；修改后点击“生成曲线”重新计算SRC。")
        form.addRow("矿点命中距离", self.src_distance_edit)
        self.src_use_test_area_check = QCheckBox("基于独立测试集绘制SRC（只用测试区域和测试矿点）")
        self.src_use_test_area_check.setChecked(False)
        form.addRow("SRC模式", self.src_use_test_area_check)
        self.src_probability_mode_combo = QComboBox()
        for label, value in PROBABILITY_MODE_CHOICES.items():
            self.src_probability_mode_combo.addItem(label, value)
        self.src_probability_mode_combo.setCurrentIndex(
            self.src_probability_mode_combo.findData(self.src_probability_mode)
        )
        form.addRow("概率计算方式", self.src_probability_mode_combo)
        uncertainty_row = QWidget()
        uncertainty_layout = QHBoxLayout(uncertainty_row)
        uncertainty_layout.setContentsMargins(0, 0, 0, 0)
        self.src_uncertainty_band_check = QCheckBox("显示多次训练阴影带")
        self.src_uncertainty_band_check.setChecked(bool(self.src_show_uncertainty_band))
        self.src_uncertainty_band_check.stateChanged.connect(lambda _state: self._update_dec_plot())
        uncertainty_layout.addWidget(self.src_uncertainty_band_check)
        self.src_uncertainty_mode_combo = QComboBox()
        self.src_uncertainty_mode_combo.addItem("±1 标准差", "std")
        self.src_uncertainty_mode_combo.addItem("95% 置信区间", "ci95")
        self.src_uncertainty_mode_combo.setCurrentIndex(
            self.src_uncertainty_mode_combo.findData(self.src_uncertainty_mode)
        )
        self.src_uncertainty_mode_combo.currentIndexChanged.connect(lambda _index: self._update_dec_plot())
        uncertainty_layout.addWidget(self.src_uncertainty_mode_combo)
        uncertainty_layout.addStretch(1)
        form.addRow("不确定性带", uncertainty_row)
        self.src_test_area_edit = self._create_path_input(
            file_filter="测试区域文件 (*.h5 *.hdf5 *.npy *.npz *.txt *.dat *.csv *.tsv *.xlsx *.xls);;所有文件 (*.*)",
            default=self.src_test_area_path,
        )
        form.addRow("测试区域文件(可选)", self.src_test_area_edit['container'])
        self.dec_deposit_edit = self._create_path_input(
            file_filter="矿点文件 (*.xlsx *.xls *.dat *.txt *.csv *.tsv);;Excel 文件 (*.xlsx *.xls);;文本文件 (*.dat *.txt *.csv *.tsv);;所有文件 (*.*)",
            default=self.dec_deposit_path,
        )
        form.addRow("矿点文件(可选)", self.dec_deposit_edit['container'])
        tab_layout.addWidget(dec_group)

        top_row = QHBoxLayout()
        tab_layout.addLayout(top_row, 0)

        list_group = QGroupBox("可用模型")
        list_layout = QVBoxLayout(list_group)
        self.dec_model_list = QListWidget()
        self.dec_model_list.setSelectionMode(QListWidget.MultiSelection)
        list_layout.addWidget(self.dec_model_list)
        hint = QLabel("提示：可直接添加 model.pth/model.pkl 或重复训练导出的 model1.pth/model2.pkl；程序会自动匹配 run_001/run_002 下的参数、测试区和矿点文件。")
        hint.setStyleSheet("color: #666666;")
        list_layout.addWidget(hint)
        top_row.addWidget(list_group, 1)

        btn_col = QVBoxLayout()
        add_btn = QPushButton("添加文件...")
        add_btn.clicked.connect(self._add_dec_models)
        btn_col.addWidget(add_btn)
        remove_btn = QPushButton("移除选中")
        remove_btn.clicked.connect(self._remove_selected_dec_models)
        btn_col.addWidget(remove_btn)
        clear_btn = QPushButton("清空列表")
        clear_btn.clicked.connect(self._clear_dec_models)
        btn_col.addWidget(clear_btn)
        gen_btn = QPushButton("生成曲线")
        gen_btn.clicked.connect(self._generate_dec_curves)
        btn_col.addWidget(gen_btn)
        legend_btn = QPushButton("修改图例")
        legend_btn.clicked.connect(lambda: self._edit_curve_legends("src"))
        btn_col.addWidget(legend_btn)
        view_btn = QPushButton("查看大图")
        view_btn.clicked.connect(self._show_dec_fullscreen)
        btn_col.addWidget(view_btn)
        save_btn = QPushButton("保存曲线")
        save_btn.clicked.connect(self._save_dec_curve)
        btn_col.addWidget(save_btn)
        btn_col.addStretch(1)
        top_row.addLayout(btn_col)

        plot_group = QGroupBox("SRC曲线")
        plot_layout = QVBoxLayout(plot_group)
        self.dec_fig = Figure(figsize=(5, 4))
        self.dec_ax = self.dec_fig.add_subplot(111)
        self.dec_canvas = FigureCanvasQTAgg(self.dec_fig)
        plot_layout.addWidget(self.dec_canvas)
        tab_layout.addWidget(plot_group, 1)
        self._clear_dec_plot()

        self.tabs.addTab(tab, "SRC曲线")

    def _build_pa_tab(self):
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.addWidget(QLabel("选择训练模型(.pth/.pkl)、已有预测结果(.dat/.h5) 或 zone_statistics.csv，直接按概率阈值生成 P-A 曲线。"))

        pa_group = QGroupBox("P-A参数")
        form = QFormLayout(pa_group)
        form.setLabelAlignment(Qt.AlignRight)
        self.pa_h5_edit = self._create_path_input(file_filter="完整H5 (*.h5 *.hdf5);;所有文件 (*.*)", default=self.pa_h5_path)
        form.addRow("完整H5(可选)", self.pa_h5_edit["container"])
        self.pa_prediction_stride_edit = self._create_line_edit(self.pa_prediction_stride)
        self.pa_prediction_stride_edit.setPlaceholderText("留空则沿用训练步长，可填 1 提高P-A分辨率")
        self.pa_prediction_stride_edit.setToolTip("从训练模型临时生成P-A预测数据时使用；窗口大小仍沿用训练参数。")
        form.addRow("P-A预测步长", self.pa_prediction_stride_edit)
        self.pa_threshold_step_edit = self._create_line_edit(self.pa_threshold_step)
        form.addRow("阈值扫描步长", self.pa_threshold_step_edit)
        self.pa_distance_edit = self._create_line_edit(self.pa_distance_threshold)
        form.addRow("距离阈值", self.pa_distance_edit)
        self.pa_use_test_area_check = QCheckBox("基于独立测试集绘制P-A（只用测试区域和测试矿点）")
        self.pa_use_test_area_check.setChecked(False)
        form.addRow("P-A模式", self.pa_use_test_area_check)
        self.pa_probability_mode_combo = QComboBox()
        for label, value in PROBABILITY_MODE_CHOICES.items():
            self.pa_probability_mode_combo.addItem(label, value)
        self.pa_probability_mode_combo.setCurrentIndex(
            self.pa_probability_mode_combo.findData(self.pa_probability_mode)
        )
        form.addRow("概率计算方式", self.pa_probability_mode_combo)
        self.pa_test_area_edit = self._create_path_input(
            file_filter="测试区域文件 (*.h5 *.hdf5 *.npy *.npz *.txt *.dat *.csv *.tsv *.xlsx *.xls);;所有文件 (*.*)",
            default=self.pa_test_area_path,
        )
        form.addRow("测试区域文件(可选)", self.pa_test_area_edit["container"])
        self.pa_deposit_edit = self._create_path_input(
            file_filter="矿点文件 (*.xlsx *.xls *.dat *.txt *.csv *.tsv);;Excel 文件 (*.xlsx *.xls);;文本文件 (*.dat *.txt *.csv *.tsv);;所有文件 (*.*)",
            default=self.pa_deposit_path,
        )
        form.addRow("矿点文件(可选)", self.pa_deposit_edit["container"])
        tab_layout.addWidget(pa_group)

        top_row = QHBoxLayout()
        tab_layout.addLayout(top_row, 0)

        list_group = QGroupBox("可用模型")
        list_layout = QVBoxLayout(list_group)
        self.pa_model_list = QListWidget()
        self.pa_model_list.setSelectionMode(QListWidget.MultiSelection)
        list_layout.addWidget(self.pa_model_list)
        hint = QLabel("提示：可直接添加 model.pth/model.pkl 或重复训练导出的 model1.pth/model2.pkl；程序会自动匹配 run_001/run_002 下的参数、测试区和矿点文件。")
        hint.setStyleSheet("color: #666666;")
        list_layout.addWidget(hint)
        top_row.addWidget(list_group, 1)

        btn_col = QVBoxLayout()
        add_btn = QPushButton("添加文件...")
        add_btn.clicked.connect(self._add_pa_models)
        btn_col.addWidget(add_btn)
        sync_btn = QPushButton("同步SRC列表")
        sync_btn.clicked.connect(self._sync_pa_models_from_src)
        btn_col.addWidget(sync_btn)
        gen_btn = QPushButton("生成P-A曲线")
        gen_btn.clicked.connect(self._generate_pa_curves)
        btn_col.addWidget(gen_btn)
        legend_btn = QPushButton("修改图例")
        legend_btn.clicked.connect(lambda: self._edit_curve_legends("pa"))
        btn_col.addWidget(legend_btn)
        view_btn = QPushButton("查看大图")
        view_btn.clicked.connect(self._show_pa_fullscreen)
        btn_col.addWidget(view_btn)
        save_btn = QPushButton("保存曲线")
        save_btn.clicked.connect(self._save_pa_curve)
        btn_col.addWidget(save_btn)
        btn_col.addStretch(1)
        top_row.addLayout(btn_col)

        result_row = QHBoxLayout()
        tab_layout.addLayout(result_row, 1)

        plot_group = QGroupBox("P-A曲线")
        plot_layout = QVBoxLayout(plot_group)
        pa_view_row = QHBoxLayout()
        pa_view_row.addWidget(QLabel("显示方式"))
        self.pa_display_mode_combo = QComboBox()
        self.pa_display_mode_combo.addItem("并排显示", "grid")
        self.pa_display_mode_combo.addItem("单个选择", "single")
        self.pa_display_mode_combo.currentIndexChanged.connect(lambda _index: self._update_pa_plot())
        pa_view_row.addWidget(self.pa_display_mode_combo)
        pa_view_row.addWidget(QLabel("模型"))
        self.pa_model_selector_combo = QComboBox()
        self.pa_model_selector_combo.setEnabled(False)
        self.pa_model_selector_combo.currentIndexChanged.connect(lambda _index: self._update_pa_plot())
        pa_view_row.addWidget(self.pa_model_selector_combo, 1)
        pa_view_row.addStretch(1)
        plot_layout.addLayout(pa_view_row)
        self.pa_fig = Figure(figsize=(5.5, 4))
        self.pa_ax = self.pa_fig.add_subplot(111)
        self.pa_canvas = FigureCanvasQTAgg(self.pa_fig)
        plot_layout.addWidget(self.pa_canvas)
        result_row.addWidget(plot_group, 2)

        table_group = QGroupBox("P-A交点")
        table_layout = QVBoxLayout(table_group)
        self.pa_summary_table = QTableWidget(0, 6)
        self.pa_summary_table.setHorizontalHeaderLabels(["模型", "交点阈值", "SR", "PAF", "交点Y", "类型"])
        self.pa_summary_table.horizontalHeader().setStretchLastSection(True)
        self.pa_summary_table.verticalHeader().setVisible(False)
        self.pa_summary_table.setEditTriggers(QTableWidget.NoEditTriggers)
        table_layout.addWidget(self.pa_summary_table)
        result_row.addWidget(table_group, 1)

        self._clear_pa_plot()
        self._update_pa_summary_table([])

        self.tabs.addTab(tab, "P-A曲线")

    def _build_shap_tab(self):
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.addWidget(QLabel("为深度学习模型生成 SHAP 可解释性分析，包含通道重要性与 summary 图。"))

        form_group = QGroupBox("参数设置")
        grid = QFormLayout(form_group)
        grid.setLabelAlignment(Qt.AlignRight)

        self.shap_model_edit = self._create_path_input(
            form_group,
            self.shap_model_path,
            file_filter="模型文件 (*.pth *.pt);;所有文件 (*.*)",
            multi_files=True,
        )
        grid.addRow("模型文件(可多选)", self.shap_model_edit['container'])

        self.shap_input_edit = self._create_path_input(form_group, self.shap_input_path, file_filter="H5 数据 (*.h5 *.hdf5);;所有文件 (*.*)")
        grid.addRow("完整H5/数据文件", self.shap_input_edit['container'])

        self.shap_label_edit = self._create_path_input(form_group, self.shap_label_path, file_filter="标签/矿点文件 (*.h5 *.hdf5 *.txt *.dat *.csv *.tsv);;所有文件 (*.*)")
        grid.addRow("标签/矿点文件", self.shap_label_edit['container'])

        self.shap_test_area_edit = self._create_path_input(form_group, self.shap_test_area_path, file_filter="测试区域 (*.h5 *.hdf5 *.npy *.npz *.txt *.dat *.csv *.tsv);;所有文件 (*.*)")
        grid.addRow("测试区域文件", self.shap_test_area_edit['container'])

        self.shap_output_edit = self._create_path_input(form_group, self.shap_output_dir, directory=True)
        grid.addRow("结果目录", self.shap_output_edit['container'])

        self.shap_background_edit = self._create_line_edit(self.shap_background)
        grid.addRow("背景样本数", self.shap_background_edit)

        self.shap_test_edit = self._create_line_edit(self.shap_test)
        grid.addRow("测试样本数", self.shap_test_edit)

        self.shap_use_test_area_check = QCheckBox("仅从独立测试集抽取背景样本和解释样本")
        self.shap_use_test_area_check.setChecked(True)
        grid.addRow("测试集模式", self.shap_use_test_area_check)

        self.shap_show_error_bars_check = QCheckBox("显示多模型重要性标准差横线")
        self.shap_show_error_bars_check.setChecked(bool(self.shap_show_error_bars))
        grid.addRow("汇总图误差线", self.shap_show_error_bars_check)

        self.shap_prior_edit = self._create_line_edit(self.shap_prior)
        grid.addRow("先验概率", self.shap_prior_edit)

        self.shap_prediction_stride_edit = self._create_line_edit(self.shap_prediction_stride)
        self.shap_prediction_stride_edit.setPlaceholderText("留空则沿用训练步长")
        self.shap_prediction_stride_edit.setToolTip("SHAP构造解释样本时的切窗步长；窗口大小仍沿用训练参数。")
        grid.addRow("SHAP切窗步长", self.shap_prediction_stride_edit)

        type_row = QWidget()
        type_layout = QHBoxLayout(type_row)
        type_layout.setContentsMargins(0, 0, 0, 0)
        label_type = QLabel("模型类型")
        label_type.setFixedWidth(80)
        type_layout.addWidget(label_type)
        self.shap_model_combo = QComboBox()
        self.shap_model_combo.addItems(get_model_display_names(SHAP_MODEL_KEYS))
        self.shap_model_combo.setCurrentText(get_model_display_name(self.shap_model_type))
        type_layout.addWidget(self.shap_model_combo)
        grid.addRow(type_row)

        device_row = QWidget()
        device_layout = QHBoxLayout(device_row)
        device_layout.setContentsMargins(0, 0, 0, 0)
        device_label = QLabel("计算设备")
        device_label.setFixedWidth(80)
        device_layout.addWidget(device_label)
        self.shap_device_combo = QComboBox()
        self.shap_device_combo.addItems(["cuda", "cpu"])
        self.shap_device_combo.setCurrentText(self.shap_device)
        device_layout.addWidget(self.shap_device_combo)
        grid.addRow(device_row)

        tab_layout.addWidget(form_group)

        run_btn = QPushButton("生成SHAP解释")
        run_btn.clicked.connect(self._run_shap)
        tab_layout.addWidget(run_btn)

        tip_label = QLabel("提示：模型文件可多选；多选时会逐个模型计算SHAP，并汇总通道重要性均值/标准差/中位数。运行日志输出在“模型评估”页签的日志窗口内。")
        tip_label.setStyleSheet("color: #666666;")
        tab_layout.addWidget(tip_label)

        status_group = QGroupBox("结果输出")
        status_layout = QVBoxLayout(status_group)
        self.shap_status_label = QLabel(self.shap_output_text)
        status_layout.addWidget(self.shap_status_label)
        tab_layout.addWidget(status_group)

        plots_group = QGroupBox("可视化结果")
        plots_layout = QHBoxLayout(plots_group)

        self.shap_channel_label = QLabel("尚未生成图像")
        self.shap_channel_label.setAlignment(Qt.AlignCenter)
        self.shap_channel_label.setCursor(Qt.PointingHandCursor)
        self.shap_channel_label.setToolTip("点击编辑该SHAP图件注释")
        self.shap_channel_label.mousePressEvent = lambda event: self._edit_shap_plot_annotations("channel")
        channel_column = QVBoxLayout()
        channel_column.addWidget(self.shap_channel_label, 1)
        channel_btn = QPushButton("查看大图")
        channel_btn.clicked.connect(lambda: self._open_shap_plot("channel"))
        channel_column.addWidget(channel_btn)
        plots_layout.addLayout(channel_column)

        self.shap_summary_label = QLabel("尚未生成图像")
        self.shap_summary_label.setAlignment(Qt.AlignCenter)
        self.shap_summary_label.setCursor(Qt.PointingHandCursor)
        self.shap_summary_label.setToolTip("点击编辑该SHAP图件注释")
        self.shap_summary_label.mousePressEvent = lambda event: self._edit_shap_plot_annotations("summary")
        summary_column = QVBoxLayout()
        summary_column.addWidget(self.shap_summary_label, 1)
        summary_btn = QPushButton("查看大图")
        summary_btn.clicked.connect(lambda: self._open_shap_plot("summary"))
        summary_column.addWidget(summary_btn)
        plots_layout.addLayout(summary_column)

        tab_layout.addWidget(plots_group, 1)

        self.tabs.addTab(tab, "SHAP解释")

    def _create_line_edit(self, text=""):
        edit = QLineEdit(str(text) if text is not None else "")
        edit.setClearButtonEnabled(True)
        return edit

    def _create_path_input(self, parent=None, default="", file_filter="所有文件 (*.*)", directory=False, multi_files=False):
        container = QWidget(parent)
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        initial_text = str(default) if default is not None else ""
        edit = QLineEdit(initial_text)
        layout.addWidget(edit, 1)

        button = QPushButton("浏览...")

        def browse():
            current_text = edit.text().strip()
            if multi_files and ";" in current_text:
                current_text = current_text.split(";", 1)[0].strip()
            start_dir = current_text or initial_text or BASE_DIR
            if directory:
                path = QFileDialog.getExistingDirectory(self, "选择目录", start_dir or BASE_DIR)
            elif multi_files:
                paths, _ = QFileDialog.getOpenFileNames(self, "选择文件", start_dir or BASE_DIR, file_filter)
                path = "; ".join(paths)
            else:
                path, _ = QFileDialog.getOpenFileName(self, "选择文件", start_dir or BASE_DIR, file_filter)
            if path:
                edit.setText(path)

        button.clicked.connect(browse)
        layout.addWidget(button)

        return {"container": container, "edit": edit, "button": button}

    def _discover_gdr_folder_models(self, root_dir: str):
        root_path = Path(root_dir)
        excluded_names = {"normalization_params.pth"}
        entries = []
        candidate_paths = (
            path
            for path in root_path.rglob("*")
            if path.suffix.lower() in MODEL_FILE_SUFFIXES
        )
        for model_path in sorted(candidate_paths, key=lambda p: str(p).lower()):
            if not model_path.is_file() or model_path.name.lower() in excluded_names:
                continue
            try:
                relative_parent = model_path.parent.relative_to(root_path)
            except ValueError:
                relative_parent = Path(model_path.parent.name)
            parts = relative_parent.parts
            folder_name = parts[0] if parts else model_path.parent.name
            training_dir = model_path.parent.name
            display_name = folder_name if training_dir == folder_name else f"{folder_name}/{training_dir}"
            entries.append(
                {
                    "path": str(model_path),
                    "file_name": model_path.name,
                    "model_name": model_path.stem,
                    "folder_name": folder_name,
                    "training_dir": training_dir,
                    "training_path": str(model_path.parent),
                    "display_name": display_name,
                }
            )
        return entries

    def _scan_gdr_model_folders(self):
        root_dir = self.gdr_model_root_edit["edit"].text().strip()
        if not root_dir:
            QMessageBox.warning(self, "缺少参数", "请先选择模型根目录")
            return []
        if not os.path.isdir(root_dir):
            QMessageBox.critical(self, "路径不存在", f"模型根目录不存在: {root_dir}")
            return []

        entries = self._discover_gdr_folder_models(root_dir)
        self.gdr_folder_model_entries = entries
        if not entries:
            self.gdr_folder_status_label.setText("未找到模型文件")
            self.log(f"[文件夹批量] 未在目录中找到可用 pth/pkl: {root_dir}")
            return []

        name_counts = {}
        for entry in entries:
            name_counts[entry["file_name"]] = name_counts.get(entry["file_name"], 0) + 1
        self.gdr_available_model_name_counts = name_counts
        preferred = PREFERRED_MODEL_FILE_NAMES
        names = sorted(name_counts, key=lambda name: (preferred.index(name) if name in preferred else len(preferred), name.lower()))
        for label, fold_names in reversed(list(FIVE_FOLD_MODEL_LABELS.items())):
            if all(name in name_counts for name in fold_names):
                names.insert(0, label)
        current_text = self.gdr_model_name_combo.currentText().strip()
        self.gdr_model_name_combo.clear()
        self.gdr_model_name_combo.addItems(names)
        if current_text and (current_text in names or ";" in current_text or "；" in current_text):
            self.gdr_model_name_combo.setCurrentText(current_text)

        summary_names = [name for name in names if name not in FIVE_FOLD_MODEL_LABELS]
        summary = "；".join(f"{name}: {name_counts[name]}个" for name in summary_names[:8])
        if len(names) > 8:
            summary += "；..."
        self.gdr_folder_status_label.setText(f"找到 {len(entries)} 个模型文件，{len(name_counts)} 类文件名")
        self.log(f"[文件夹批量] 扫描完成: {root_dir}")
        self.log(f"[文件夹批量] {summary}")
        return entries

    def _available_gdr_model_names(self):
        name_counts = dict(getattr(self, "gdr_available_model_name_counts", {}) or {})
        if not name_counts and self.gdr_folder_model_entries:
            for entry in self.gdr_folder_model_entries:
                file_name = entry.get("file_name", "")
                if file_name:
                    name_counts[file_name] = name_counts.get(file_name, 0) + 1
        preferred = PREFERRED_MODEL_FILE_NAMES
        return sorted(name_counts, key=lambda name: (preferred.index(name) if name in preferred else len(preferred), name.lower()))

    def _selected_gdr_model_names_from_text(self):
        text = self.gdr_model_name_combo.currentText().strip()
        if not text:
            return []
        if text in FIVE_FOLD_MODEL_LABELS:
            return list(FIVE_FOLD_MODEL_LABELS[text])
        normalized = text.replace("；", ";").replace(",", ";").replace("\n", ";")
        names = [part.strip() for part in normalized.split(";") if part.strip()]
        return names or [text]

    def _selected_gdr_model_names_are_complete_group(self, selected_model_names):
        selected = [str(name).strip() for name in selected_model_names if str(name).strip()]
        return selected in (FOLD_MODEL_NAMES, PKL_FOLD_MODEL_NAMES)

    def _select_gdr_model_names(self):
        if not self.gdr_folder_model_entries:
            self._scan_gdr_model_folders()
        names = self._available_gdr_model_names()
        if not names:
            QMessageBox.information(self, "提示", "请先扫描包含模型文件的根目录")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("选择模型文件名")
        dialog.resize(420, 360)
        layout = QVBoxLayout(dialog)
        hint = QLabel("选择需要纳入批量评估的模型文件名，可多选。")
        layout.addWidget(hint)

        list_widget = QListWidget()
        list_widget.setSelectionMode(QListWidget.MultiSelection)
        current_names = set(self._selected_gdr_model_names_from_text())
        if not current_names:
            current_names = set(FOLD_MODEL_NAMES)
        for name in names:
            list_widget.addItem(f"{name} ({self.gdr_available_model_name_counts.get(name, 0)}个)")
            item = list_widget.item(list_widget.count() - 1)
            item.setData(Qt.UserRole, name)
            if name in current_names:
                item.setSelected(True)
        layout.addWidget(list_widget, 1)

        btn_row = QHBoxLayout()
        select_folds_btn = QPushButton("选择fold1-5")
        def select_folds():
            available_names = {
                list_widget.item(row).data(Qt.UserRole)
                for row in range(list_widget.count())
            }
            fold_names = set(FOLD_MODEL_NAMES)
            if not fold_names.issubset(available_names) and set(PKL_FOLD_MODEL_NAMES).issubset(available_names):
                fold_names = set(PKL_FOLD_MODEL_NAMES)
            for row in range(list_widget.count()):
                item = list_widget.item(row)
                item.setSelected(item.data(Qt.UserRole) in fold_names)
        select_folds_btn.clicked.connect(select_folds)
        btn_row.addWidget(select_folds_btn)
        btn_row.addStretch(1)
        ok_btn = QPushButton("确定")
        cancel_btn = QPushButton("取消")
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)

        if dialog.exec_() != QDialog.Accepted:
            return
        selected_names = [
            list_widget.item(row).data(Qt.UserRole)
            for row in range(list_widget.count())
            if list_widget.item(row).isSelected()
        ]
        if not selected_names:
            QMessageBox.warning(self, "未选择文件名", "请至少选择一个模型文件名")
            return
        self.gdr_model_name_combo.setCurrentText("; ".join(selected_names))
        self.gdr_folder_status_label.setText(f"已选择 {len(selected_names)} 类模型文件")
        self.log(f"[文件夹批量] 已选择模型文件名: {'; '.join(selected_names)}")

    def _selected_gdr_folder_groups(self):
        if not self.gdr_folder_model_entries:
            self._scan_gdr_model_folders()
        if not self.gdr_folder_model_entries:
            return []
        selected_model_names = self._selected_gdr_model_names_from_text()
        if not selected_model_names:
            QMessageBox.warning(self, "缺少参数", "请填写或选择模型文件名")
            return []

        groups = {}
        for entry in self.gdr_folder_model_entries:
            if entry.get("file_name") not in selected_model_names:
                continue
            groups.setdefault(entry.get("training_path"), []).append(entry)

        complete_groups = []
        fold_order = {name: index for index, name in enumerate(selected_model_names)}
        for training_path, entries in groups.items():
            by_name = {entry["file_name"]: entry for entry in entries}
            if not all(name in by_name for name in selected_model_names):
                continue
            ordered_entries = [by_name[name] for name in selected_model_names]
            ordered_entries = sorted(ordered_entries, key=lambda item: fold_order.get(item["file_name"], 99))
            first = ordered_entries[0]
            complete_groups.append(
                {
                    "entries": ordered_entries,
                    "folder_name": first.get("folder_name", ""),
                    "training_dir": first.get("training_dir", ""),
                    "training_path": training_path,
                    "display_name": first.get("display_name") or first.get("folder_name") or first.get("training_dir"),
                }
            )

        complete_groups = sorted(complete_groups, key=lambda item: str(item.get("training_path", "")).lower())
        if not complete_groups:
            QMessageBox.warning(self, "未找到完整模型组", "未找到同时包含所选模型文件名的子文件夹")
        return complete_groups

    def _selected_gdr_folder_entries(self):
        root_dir = self.gdr_model_root_edit["edit"].text().strip()
        if not self.gdr_folder_model_entries:
            self._scan_gdr_model_folders()
        if not self.gdr_folder_model_entries:
            return []
        selected_name = self.gdr_model_name_combo.currentText().strip()
        if not selected_name:
            QMessageBox.warning(self, "缺少参数", "请填写或选择模型文件名")
            return []
        selected_model_names = self._selected_gdr_model_names_from_text()
        if self._selected_gdr_model_names_are_complete_group(selected_model_names):
            groups = self._selected_gdr_folder_groups()
            return [entry for group in groups for entry in group["entries"]]
        selected_name_set = {name.lower() for name in selected_model_names}
        selected_entries = [
            entry for entry in self.gdr_folder_model_entries
            if entry.get("file_name", "").lower() in selected_name_set
        ]
        if not selected_entries:
            QMessageBox.warning(self, "未找到模型", f"在 {root_dir} 下没有找到所选模型文件名")
            return []
        selected_entries = sorted(selected_entries, key=lambda item: str(item.get("path", "")).lower())
        return selected_entries

    def _apply_gdr_folder_models(self):
        selected_entries = self._selected_gdr_folder_entries()
        if not selected_entries:
            return
        self.eval_model_edit["edit"].setText("; ".join(entry["path"] for entry in selected_entries))
        selected_model_names = self._selected_gdr_model_names_from_text()
        if self._selected_gdr_model_names_are_complete_group(selected_model_names):
            groups = self._selected_gdr_folder_groups()
            self.gdr_folder_status_label.setText(f"已选择 {len(groups)} 组，共 {len(selected_entries)} 个模型")
            self.log(f"[文件夹批量] 已写入 {len(groups)} 组模型到“模型文件”输入框: {'; '.join(selected_model_names)}")
        else:
            self.gdr_folder_status_label.setText(f"已选择 {len(selected_entries)} 个模型")
            self.log(f"[文件夹批量] 已写入 {len(selected_entries)} 个模型到“模型文件”输入框")

    def _run_gdr_folder_batch_csv(self):
        selected_model_names = self._selected_gdr_model_names_from_text()
        if self._selected_gdr_model_names_are_complete_group(selected_model_names):
            model_groups = self._selected_gdr_folder_groups()
            if not model_groups:
                return
            selected_entries = [entry for group in model_groups for entry in group["entries"]]
        else:
            model_groups = []
            selected_entries = self._selected_gdr_folder_entries()
        if not selected_entries:
            return
        self.eval_model_edit["edit"].setText("; ".join(entry["path"] for entry in selected_entries))
        config = self._collect_gdr_config()
        if not config:
            return

        default_suffix = "xlsx" if model_groups else "csv"
        default_output = os.path.join(
            self.gdr_model_root_edit["edit"].text().strip() or BASE_DIR,
            f"模型评估结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{default_suffix}",
        )
        file_filter = "Excel 文件 (*.xlsx);;CSV 文件 (*.csv);;所有文件 (*.*)" if model_groups else "CSV 文件 (*.csv);;所有文件 (*.*)"
        output_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存批量评估结果",
            default_output,
            file_filter,
        )
        if not output_path:
            return
        if Path(output_path).suffix.lower() not in {".csv", ".xlsx"}:
            output_path = str(Path(output_path).with_suffix(f".{default_suffix}"))

        config["model_entries"] = selected_entries
        config["model_paths"] = [entry["path"] for entry in selected_entries]
        config["model_path"] = selected_entries[0]["path"]
        config["csv_output_path"] = output_path
        config["use_model_metadata_per_folder"] = True
        if model_groups:
            config["model_groups"] = model_groups
            config["group_model_names"] = selected_model_names
            target = run_gdr_grouped_folds_task
            success_message = f"分组批量评估完成，结果已保存至: {output_path}"
        else:
            target = run_gdr_batch_task
            success_message = f"模型评估完成，结果已保存至: {output_path}"
        self._run_async(
            target,
            args=(config, self.log),
            success_message=success_message,
            result_handler=self._handle_evaluation_result,
        )

    def _run_gdr(self):
        config = self._collect_gdr_config()
        if not config:
            return
        model_paths = list(config.get("model_paths") or [])
        target = run_gdr_batch_task if len(model_paths) > 1 else run_gdr_task
        self._run_async(
            target,
            args=(config, self.log),
            success_message="模型评估完成",
            result_handler=self._handle_evaluation_result,
        )

    def _collect_gdr_config(self):
        prediction_file = self.prediction_edit['edit'].text().strip()
        model_text = self.eval_model_edit['edit'].text().strip()
        normalized_model_text = model_text.replace("；", ";").replace("\n", ";")
        model_paths = [part.strip().strip('"') for part in normalized_model_text.split(";") if part.strip()]
        model_path = model_paths[0] if model_paths else ""
        data_path = self.eval_data_edit['edit'].text().strip()
        deposit_file = self.gdr_deposit_edit['edit'].text().strip()
        independent_test = self.independent_test_check.isChecked()
        test_area_file = self.test_area_edit['edit'].text().strip()
        training_meta = {}
        model_dir = os.path.dirname(model_path) if model_path else ""
        if model_path:
            try:
                model_dir, training_meta = _load_training_metadata(model_path)
            except Exception:
                training_meta = {}
        if model_path and not data_path:
            data_path = _existing_path(training_meta.get("dataset"))
        if model_path and independent_test and not test_area_file:
            test_area_file = _first_existing_path(training_meta.get("test_area_path"), os.path.join(model_dir, "test_area.h5"))
        if model_path and not deposit_file:
            if independent_test:
                deposit_file = _first_existing_path(
                    training_meta.get("test_mineral_path"),
                    os.path.join(model_dir, "test_minerals.txt"),
                    training_meta.get("label_path"),
                )
            else:
                deposit_file = _first_existing_path(training_meta.get("label_path"), os.path.join(model_dir, "test_minerals.txt"))
        missing = []
        if not prediction_file and not model_paths:
            missing.append("预测文件或模型文件")
        if model_paths and not data_path:
            missing.append("完整特征H5")
        if not deposit_file:
            missing.append("矿点文件")
        if missing:
            QMessageBox.warning(self, "缺少参数", f"请填写: {', '.join(missing)}")
            return None
        paths_to_check = [("矿点文件", deposit_file)]
        if prediction_file and not model_paths:
            paths_to_check.append(("预测文件", prediction_file))
        if model_paths:
            for index, item_path in enumerate(model_paths, start=1):
                paths_to_check.append((f"模型文件 {index}", item_path))
            paths_to_check.append(("完整特征H5", data_path))
        if test_area_file:
            paths_to_check.append(("测试区域文件", test_area_file))
        for label, path in paths_to_check:
            if not os.path.exists(path):
                QMessageBox.critical(self, "路径不存在", f"{label} 不存在: {path}")
                return None

        try:
            distance_threshold = float(self.distance_edit.text())
            confidence_threshold = float(self.confidence_edit.text())
            prior = float(self.prior_edit.text())
            threshold_step = float(self.threshold_step_edit.text())
            prediction_patch_stride = _optional_positive_int(self.gdr_prediction_stride_edit.text(), "评估步长")
        except ValueError:
            QMessageBox.critical(self, "参数错误", "请确保阈值为数值类型，评估步长为正整数或留空")
            return None
        if threshold_step <= 0 or threshold_step > 1:
            QMessageBox.critical(self, "参数错误", "阈值扫描步长需要在 0-1 之间")
            return None

        return {
            "prediction_file": prediction_file,
            "model_path": model_path,
            "model_paths": model_paths,
            "data_path": data_path,
            "label_file": deposit_file,
            "deposit_file": deposit_file,
            "test_area_file": test_area_file,
            "independent_test": independent_test,
            "threshold_strategy": "fixed" if self.threshold_strategy_combo.currentText() == "手动阈值" else "max_ei",
            "distance_threshold": distance_threshold,
            "confidence_threshold": confidence_threshold,
            "threshold_step": threshold_step,
            "prior": prior,
            "probability_mode": self.gdr_probability_mode_combo.currentData() or PROBABILITY_MODE_CENTER,
            "prediction_patch_stride": prediction_patch_stride,
        }

    def _run_shap(self):
        config = self._collect_shap_config()
        if not config:
            return
        self.shap_status_label.setText("正在运行 SHAP 分析，请查看日志...")
        self._run_async(
            run_shap_task,
            args=(config, self.log),
            success_message="SHAP分析完成",
            result_handler=self._handle_shap_result,
        )

    def _collect_shap_config(self):
        raw_model_text = self.shap_model_edit['edit'].text().strip()
        model_paths = [part.strip().strip('"') for part in raw_model_text.replace("\n", ";").split(";") if part.strip()]
        model_path = model_paths[0] if model_paths else ""
        input_path = self.shap_input_edit['edit'].text().strip()
        label_path = self.shap_label_edit['edit'].text().strip()
        test_area_path = self.shap_test_area_edit['edit'].text().strip()
        output_dir = self.shap_output_edit['edit'].text().strip()
        use_test_area = self.shap_use_test_area_check.isChecked()

        model_dir, training_meta = ("", {})
        if model_path:
            try:
                model_dir, training_meta = _load_training_metadata(model_path)
            except Exception:
                model_dir, training_meta = (os.path.dirname(model_path), {})

        if model_path and (not input_path or not os.path.exists(input_path)):
            input_path = str(training_meta.get("dataset") or "").strip()
            if input_path:
                self.shap_input_edit['edit'].setText(input_path)
        if model_path and (not label_path or not os.path.exists(label_path)):
            label_path = str(training_meta.get("label_path") or "").strip()
            if label_path:
                self.shap_label_edit['edit'].setText(label_path)
        if model_path and (not test_area_path or not os.path.exists(test_area_path)):
            test_area_path = str(training_meta.get("test_area_path") or os.path.join(model_dir, "test_area.h5")).strip()
            if test_area_path and os.path.exists(test_area_path):
                self.shap_test_area_edit['edit'].setText(test_area_path)

        for path in model_paths:
            if not os.path.exists(path):
                QMessageBox.critical(self, "路径不存在", f"模型文件不存在: {path}")
                return None
        for label, path in ("模型文件", model_path), ("完整H5/数据文件", input_path), ("标签/矿点文件", label_path):
            if not path:
                QMessageBox.warning(self, "缺少参数", f"请填写 {label} 路径")
                return None
            if not os.path.exists(path):
                QMessageBox.critical(self, "路径不存在", f"{label} 不存在: {path}")
                return None
        if use_test_area:
            if not test_area_path:
                QMessageBox.warning(self, "缺少参数", "已启用测试集模式，请填写测试区域文件")
                return None
            if not os.path.exists(test_area_path):
                QMessageBox.critical(self, "路径不存在", f"测试区域文件不存在: {test_area_path}")
                return None

        try:
            num_background = int(self.shap_background_edit.text())
            num_test = int(self.shap_test_edit.text())
            shap_prediction_stride = _optional_positive_int(self.shap_prediction_stride_edit.text(), "SHAP切窗步长")
        except ValueError:
            QMessageBox.critical(self, "参数错误", "请确保样本数为整数，SHAP切窗步长为正整数或留空")
            return None
        try:
            prior = float(
                training_meta.get("resolved_prior")
                if training_meta.get("resolved_prior") is not None
                else training_meta.get("calculated_prior")
                if training_meta.get("calculated_prior") is not None
                else self.shap_prior_edit.text()
            )
        except ValueError:
            QMessageBox.critical(self, "参数错误", "请确保先验为数值类型")
            return None
        self.shap_prior_edit.setText(str(prior))

        if num_background <= 0 or num_test <= 0:
            QMessageBox.critical(self, "参数错误", "样本数需为正整数")
            return None

        model_type = normalize_model_key(training_meta.get("model") or self.shap_model_combo.currentText())
        device = self.shap_device_combo.currentText().lower()
        norm_params_path = str(training_meta.get("normalization_params_path") or "").strip()
        if not norm_params_path or not os.path.exists(norm_params_path):
            norm_params_path = os.path.join(model_dir, "normalization_params.pth")
        patch_size = int(training_meta.get("patch_size") or training_meta.get("img_size") or 16)
        training_patch_stride = int(training_meta.get("patch_stride") or patch_size)
        patch_stride = int(shap_prediction_stride or training_patch_stride)
        reflect_padding = bool(training_meta.get("reflect_padding", True))
        selected_channels = str(training_meta.get("selected_channels") or "")
        for other_model_path in model_paths[1:]:
            _, other_meta = _load_training_metadata(other_model_path)
            other_model_type = normalize_model_key(other_meta.get("model") or model_type)
            other_patch_size = int(other_meta.get("patch_size") or other_meta.get("img_size") or patch_size)
            other_channels = str(other_meta.get("selected_channels") or "")
            if other_model_type != model_type:
                QMessageBox.critical(
                    self,
                    "模型不一致",
                    f"多模型SHAP要求模型类型一致: {os.path.basename(other_model_path)} 为 {get_model_display_name(other_model_type)}，首个模型为 {get_model_display_name(model_type)}",
                )
                return None
            if other_patch_size != patch_size:
                QMessageBox.critical(
                    self,
                    "窗口大小不一致",
                    f"多模型SHAP要求窗口大小一致: {os.path.basename(other_model_path)} 为 {other_patch_size}，首个模型为 {patch_size}",
                )
                return None
            if other_channels != selected_channels:
                QMessageBox.critical(
                    self,
                    "图层不一致",
                    f"多模型SHAP要求 selected_channels 一致: {os.path.basename(other_model_path)} 与首个模型不同",
                )
                return None

        if output_dir:
            try:
                os.makedirs(output_dir, exist_ok=True)
            except OSError as exc:
                QMessageBox.critical(self, "路径错误", f"无法创建结果目录: {exc}")
                return None

        return {
            "model_path": model_path,
            "model_paths": model_paths,
            "model_type": model_type,
            "input_data": input_path,
            "label_file": label_path,
            "test_area_file": test_area_path,
            "use_test_area": use_test_area,
            "use_prediction_dataset": True,
            "norm_params_path": norm_params_path,
            "patch_size": patch_size,
            "patch_stride": patch_stride,
            "reflect_padding": reflect_padding,
            "selected_channels": selected_channels,
            "prior": prior,
            "num_background_samples": num_background,
            "num_test_samples": num_test,
            "device": device,
            "show_error_bars": self.shap_show_error_bars_check.isChecked(),
            "custom_channel_names": "",
            "importance_xlabel": self.shap_importance_xlabel,
            "summary_xlabel": self.shap_summary_xlabel,
            "figure_dpi": 300,
            "channel_importance_figsize": (12, 6),
            "summary_plot_figsize": (10, 8),
            "output_dir": output_dir,
        }


    def _add_dec_models(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "选择SRC曲线数据文件",
            BASE_DIR,
            "支持的文件 (*.pth *.pkl *.dat *.h5 *.hdf5 *.json *.csv);;模型文件 (*.pth *.pkl);;预测文件 (*.dat *.h5 *.hdf5);;JSON Files (*.json);;CSV Files (*.csv);;所有文件 (*.*)",
        )
        added = 0
        for file_path in files or []:
            path_obj = Path(file_path)
            suffix = path_obj.suffix.lower()
            entries = []

            if suffix in MODEL_FILE_SUFFIXES:
                try:
                    model_dir, training_meta = _load_training_metadata(file_path)
                    model_key = normalize_model_key(training_meta.get("model") or path_obj.stem)
                    model_name = get_model_display_name(model_key) if model_key else path_obj.stem
                    distance_threshold = float(training_meta.get("spatial_metric_distance_threshold") or 4.0)
                    entries.append(
                        {
                            "name": model_name,
                            "path": file_path,
                            "disambiguation": path_obj.stem,
                            "source_type": "trained_model",
                            "data_path": str(training_meta.get("dataset") or "").strip(),
                            "test_area_path": _first_existing_path(
                                training_meta.get("test_area_path"),
                                os.path.join(model_dir, "test_area.h5"),
                            ),
                            "test_mineral_path": _first_existing_path(
                                training_meta.get("test_mineral_path"),
                                os.path.join(model_dir, "test_minerals.txt"),
                            ),
                            "distance_threshold": distance_threshold,
                        }
                    )
                except Exception as exc:
                    self.log(f"[SRC] 读取模型参数失败: {exc}")
                    continue
            elif path_obj.name.lower() == "model_packages.json":
                try:
                    with open(file_path, "r", encoding="utf-8") as fh:
                        packages = json.load(fh)
                    for item in packages or []:
                        manifest_path = str(item.get("rebuild_manifest_path") or "").strip()
                        model_name = str(item.get("model_name") or Path(manifest_path).parent.name or path_obj.stem)
                        if manifest_path:
                            entries.append({"name": model_name, "path": manifest_path, "source_type": "manifest"})
                except Exception as exc:
                    self.log(f"[SRC] 读取 model_packages.json 失败: {exc}")
                    continue
            elif path_obj.name.lower() == "rebuild_manifest.json" or suffix == ".json":
                try:
                    with open(file_path, "r", encoding="utf-8") as fh:
                        manifest = json.load(fh)
                    model_name = str(manifest.get("model_name") or path_obj.parent.name or path_obj.stem)
                    if "zone_statistics_path" in manifest or "prediction_header_path" in manifest or "model_name" in manifest:
                        entries.append({"name": model_name, "path": file_path, "source_type": "manifest"})
                except Exception as exc:
                    self.log(f"[SRC] 读取清单文件失败: {exc}")
                    continue
            elif path_obj.name.lower() == "zone_statistics.csv" or suffix == ".csv":
                model_name = path_obj.parent.name if path_obj.parent.name else path_obj.stem
                entries.append({"name": model_name, "path": file_path, "source_type": "zone_csv"})
            else:
                raw_name = path_obj.stem
                display_name = get_model_display_name(normalize_model_key(raw_name))
                source_type = "prediction_h5" if suffix in {".h5", ".hdf5"} else "dat"
                if source_type == "prediction_h5":
                    try:
                        with h5py.File(file_path, "r") as fh:
                            attrs = dict(fh.attrs)
                        model_type = normalize_model_key(attrs.get("prediction_model_type"))
                        model_name = get_model_display_name(model_type) if model_type else display_name
                        method = str(attrs.get("prediction_aggregation_method") or "").strip()
                        count = attrs.get("prediction_model_count")
                        if method and count:
                            model_name = f"{model_name}_{method}{int(count)}"
                        display_name = model_name
                    except Exception as exc:
                        self.log(f"[SRC] 读取预测H5元数据失败，使用文件名显示: {exc}")
                entries.append({"name": display_name, "path": file_path, "source_type": source_type})

            for entry in entries:
                if not any(
                    model["path"] == entry["path"] and str(model.get("source_type")) == str(entry.get("source_type"))
                    for model in self.dec_available_models
                ):
                    self.dec_available_models.append(entry)
                    added += 1
        if added:
            self._refresh_dec_model_list()
            self.log(f"[SRC] 已添加 {added} 个曲线数据源")

    def _dec_model_key(self, model: dict):
        path = str(model.get("path") or "")
        source_type = str(model.get("source_type") or "")
        if path or source_type:
            return (path, source_type)
        return ("", str(model.get("name") or ""))

    def _remove_dec_curves_by_keys(self, removed_keys):
        if not removed_keys or not getattr(self, "dec_last_result", None):
            return 0
        before_count = len(self.dec_last_result)
        self.dec_last_result = [
            curve
            for curve in self.dec_last_result
            if self._dec_model_key(curve) not in removed_keys
        ]
        removed_count = before_count - len(self.dec_last_result)
        if removed_count:
            self._update_dec_plot()
        return removed_count

    def _remove_selected_dec_models(self):
        indexes = sorted(
            {index.row() for index in self.dec_model_list.selectedIndexes()},
            reverse=True,
        )
        if not indexes:
            QMessageBox.information(self, "提示", "请先在SRC模型列表中选择要移除的模型。")
            return

        removed_models = []
        for row in indexes:
            if 0 <= row < len(self.dec_available_models):
                removed_models.append(self.dec_available_models.pop(row))

        removed_keys = {self._dec_model_key(model) for model in removed_models}
        removed_curve_count = self._remove_dec_curves_by_keys(removed_keys)
        self._refresh_dec_model_list()
        if not self.dec_last_result:
            self._clear_dec_plot()
        self.log(f"[SRC] 已移除 {len(removed_models)} 个模型，已同步移除 {removed_curve_count} 条曲线")

    def _clear_dec_models(self):
        if not self.dec_available_models:
            QMessageBox.information(self, "提示", "SRC模型列表已经为空。")
            return
        removed_count = len(self.dec_available_models)
        self.dec_available_models.clear()
        self.dec_model_list.clear()
        curve_count = len(getattr(self, "dec_last_result", []) or [])
        self.dec_last_result = []
        self._clear_dec_plot()
        self.log(f"[SRC] 已清空 {removed_count} 个模型，已移除 {curve_count} 条曲线")

    def _add_pa_models(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "选择P-A曲线数据文件",
            BASE_DIR,
            "支持的文件 (*.pth *.pkl *.dat *.h5 *.hdf5 *.json *.csv);;模型文件 (*.pth *.pkl);;预测文件 (*.dat *.h5 *.hdf5);;JSON Files (*.json);;CSV Files (*.csv);;所有文件 (*.*)",
        )
        added = 0
        for file_path in files or []:
            path_obj = Path(file_path)
            suffix = path_obj.suffix.lower()
            entries = []

            if suffix in MODEL_FILE_SUFFIXES:
                try:
                    model_dir, training_meta = _load_training_metadata(file_path)
                    model_key = normalize_model_key(training_meta.get("model") or path_obj.stem)
                    model_name = get_model_display_name(model_key) if model_key else path_obj.stem
                    entries.append(
                        {
                            "name": model_name,
                            "path": file_path,
                            "disambiguation": path_obj.stem,
                            "source_type": "trained_model",
                            "data_path": str(training_meta.get("dataset") or "").strip(),
                            "test_area_path": _first_existing_path(
                                training_meta.get("test_area_path"),
                                os.path.join(model_dir, "test_area.h5"),
                            ),
                            "test_mineral_path": _first_existing_path(
                                training_meta.get("test_mineral_path"),
                                os.path.join(model_dir, "test_minerals.txt"),
                            ),
                            "distance_threshold": 4.0,
                        }
                    )
                except Exception as exc:
                    self.log(f"[P-A] 读取模型参数失败: {exc}")
                    continue
            elif path_obj.name.lower() == "model_packages.json":
                try:
                    with open(file_path, "r", encoding="utf-8") as fh:
                        packages = json.load(fh)
                    for item in packages or []:
                        manifest_path = str(item.get("rebuild_manifest_path") or "").strip()
                        model_name = str(item.get("model_name") or Path(manifest_path).parent.name or path_obj.stem)
                        if manifest_path:
                            entries.append({"name": model_name, "path": manifest_path, "source_type": "manifest"})
                except Exception as exc:
                    self.log(f"[P-A] 读取 model_packages.json 失败: {exc}")
                    continue
            elif path_obj.name.lower() == "rebuild_manifest.json" or suffix == ".json":
                try:
                    with open(file_path, "r", encoding="utf-8") as fh:
                        manifest = json.load(fh)
                    model_name = str(manifest.get("model_name") or path_obj.parent.name or path_obj.stem)
                    if "zone_statistics_path" in manifest or "model_name" in manifest:
                        entries.append({"name": model_name, "path": file_path, "source_type": "manifest"})
                except Exception as exc:
                    self.log(f"[P-A] 读取清单文件失败: {exc}")
                    continue
            elif path_obj.name.lower() == "zone_statistics.csv" or suffix == ".csv":
                model_name = path_obj.parent.name if path_obj.parent.name else path_obj.stem
                entries.append({"name": model_name, "path": file_path, "source_type": "zone_csv"})
            else:
                raw_name = path_obj.stem
                display_name = get_model_display_name(normalize_model_key(raw_name))
                source_type = "prediction_h5" if suffix in {".h5", ".hdf5"} else "dat"
                entries.append({"name": display_name, "path": file_path, "source_type": source_type})

            for entry in entries:
                if not any(
                    model["path"] == entry["path"] and str(model.get("source_type")) == str(entry.get("source_type"))
                    for model in self.pa_available_models
                ):
                    self.pa_available_models.append(entry)
                    added += 1
        if added:
            self._refresh_pa_model_list()
            self.log(f"[P-A] 已添加 {added} 个曲线数据源")

    def _sync_pa_models_from_src(self):
        if not self.dec_available_models:
            QMessageBox.information(self, "提示", "SRC列表为空，请先添加文件或直接在P-A页添加。")
            return
        existing = {
            (model.get("path"), str(model.get("source_type") or ""))
            for model in self.pa_available_models
        }
        added = 0
        for model in self.dec_available_models:
            key = (model.get("path"), str(model.get("source_type") or ""))
            if key in existing:
                continue
            self.pa_available_models.append(dict(model))
            existing.add(key)
            added += 1
        self._refresh_pa_model_list()
        self.log(f"[P-A] 已从SRC列表同步 {added} 个数据源")

    def _display_dec_model_name(self, model: dict, duplicate_names: set[str]) -> str:
        name = str(model.get("name") or Path(str(model.get("path") or "")).stem)
        if normalize_model_key(name) == "linear":
            name = "Linear"
        if name in duplicate_names and model.get("disambiguation"):
            return f"{name} ({model['disambiguation']})"
        return name

    def _display_pa_model_name(self, model: dict, duplicate_names: set[str]) -> str:
        name = str(model.get("name") or Path(str(model.get("path") or "")).stem)
        if name in duplicate_names and model.get("disambiguation"):
            return f"{name} ({model['disambiguation']})"
        return name

    def _refresh_dec_model_list(self):
        selected_paths = {
            self.dec_available_models[index.row()]["path"]
            for index in self.dec_model_list.selectedIndexes()
            if 0 <= index.row() < len(self.dec_available_models)
        }
        name_counts = {}
        for model in self.dec_available_models:
            name = str(model.get("name") or "")
            name_counts[name] = name_counts.get(name, 0) + 1
        duplicate_names = {name for name, count in name_counts.items() if count > 1}

        self.dec_model_list.clear()
        for row, model in enumerate(self.dec_available_models):
            source_tag = str(model.get("source_type") or "dat").upper()
            display_name = self._display_dec_model_name(model, duplicate_names)
            self.dec_model_list.addItem(f"{display_name}  [{source_tag}]  ({model['path']})")
            if model.get("path") in selected_paths:
                self.dec_model_list.item(row).setSelected(True)

    def _get_selected_dec_models(self):
        items = self.dec_model_list.selectedIndexes()
        if not items:
            return []
        models = []
        for index in items:
            row = index.row()
            if 0 <= row < len(self.dec_available_models):
                models.append(self.dec_available_models[row])
        return models

    def _refresh_pa_model_list(self):
        selected_paths = {
            self.pa_available_models[index.row()]["path"]
            for index in self.pa_model_list.selectedIndexes()
            if 0 <= index.row() < len(self.pa_available_models)
        }
        name_counts = {}
        for model in self.pa_available_models:
            name = str(model.get("name") or "")
            name_counts[name] = name_counts.get(name, 0) + 1
        duplicate_names = {name for name, count in name_counts.items() if count > 1}

        self.pa_model_list.clear()
        for row, model in enumerate(self.pa_available_models):
            source_tag = str(model.get("source_type") or "dat").upper()
            display_name = self._display_pa_model_name(model, duplicate_names)
            self.pa_model_list.addItem(f"{display_name}  [{source_tag}]  ({model['path']})")
            if model.get("path") in selected_paths:
                self.pa_model_list.item(row).setSelected(True)

    def _get_selected_pa_models(self):
        items = self.pa_model_list.selectedIndexes()
        if not items:
            return []
        models = []
        for index in items:
            row = index.row()
            if 0 <= row < len(self.pa_available_models):
                models.append(self.pa_available_models[row])
        return models

    def _generate_dec_curves(self):
        models = self._get_selected_dec_models()
        if not models:
            QMessageBox.warning(self, "未选择模型", "请至少选择一个模型进行SRC计算")
            return
        name_counts = {}
        for model in models:
            name = str(model.get("name") or "")
            name_counts[name] = name_counts.get(name, 0) + 1
        duplicate_names = {name for name, count in name_counts.items() if count > 1}
        for model in models:
            group_name = str(model.get("name") or Path(str(model.get("path") or "")).stem)
            if normalize_model_key(group_name) == "linear":
                group_name = "Linear"
            model["curve_group_name"] = group_name
            model["plot_name"] = self._display_dec_model_name(model, duplicate_names)
        h5_path = self.src_h5_edit['edit'].text().strip()
        mineral_path = self.dec_deposit_edit['edit'].text().strip()
        use_test_area = self.src_use_test_area_check.isChecked()
        test_area_path = self.src_test_area_edit['edit'].text().strip()

        for model in models:
            if str(model.get("source_type") or "").lower() != "trained_model":
                continue
            if h5_path and (not model.get("data_path") or not os.path.exists(str(model.get("data_path")))):
                model["data_path"] = h5_path
            if test_area_path and (not model.get("test_area_path") or not os.path.exists(str(model.get("test_area_path")))):
                model["test_area_path"] = test_area_path
            if mineral_path and (
                not use_test_area
                or not model.get("test_mineral_path")
                or not os.path.exists(str(model.get("test_mineral_path")))
            ):
                model["test_mineral_path"] = mineral_path

            if not model.get("data_path"):
                QMessageBox.critical(self, "缺少参数", f"{model.get('name')} 缺少完整H5，请在SRC参数中选择完整H5文件")
                return
            if not os.path.exists(str(model.get("data_path"))):
                QMessageBox.critical(self, "路径不存在", f"完整H5不存在: {model.get('data_path')}")
                return
            if use_test_area and not model.get("test_area_path"):
                QMessageBox.critical(self, "缺少参数", f"{model.get('name')} 缺少测试区域文件，请在SRC参数中选择测试区域文件")
                return
            if use_test_area and not os.path.exists(str(model.get("test_area_path"))):
                QMessageBox.critical(self, "路径不存在", f"测试区域文件不存在: {model.get('test_area_path')}")
                return
            if not model.get("test_mineral_path"):
                QMessageBox.critical(self, "缺少参数", f"{model.get('name')} 缺少矿点文件，请在SRC参数中选择矿点文件")
                return
            if not os.path.exists(str(model.get("test_mineral_path"))):
                QMessageBox.critical(self, "路径不存在", f"矿点文件不存在: {model.get('test_mineral_path')}")
                return

        prediction_source_types = {"dat", "prediction_h5"}
        has_prediction_file_model = any(
            str(model.get("source_type") or "dat").lower() in prediction_source_types
            for model in models
        )
        requires_deposit = has_prediction_file_model
        deposit_path = mineral_path
        if requires_deposit:
            if not deposit_path:
                QMessageBox.critical(self, "缺少参数", "所选 DAT/H5 预测结果需要矿点文件参与计算，请先选择矿点文件")
                return
            if not os.path.exists(deposit_path):
                QMessageBox.critical(self, "路径不存在", f"矿点文件不存在: {deposit_path}")
                return
        if use_test_area:
            trained_models = [
                model
                for model in models
                if str(model.get("source_type") or "").lower() == "trained_model"
            ]
            trained_have_area = all(str(model.get("test_area_path") or "").strip() for model in trained_models)
            if not test_area_path and (has_prediction_file_model or not trained_have_area):
                QMessageBox.critical(self, "缺少参数", "独立测试集SRC需要测试区域文件")
                return
            if test_area_path and not os.path.exists(test_area_path):
                QMessageBox.critical(self, "路径不存在", f"测试区域文件不存在: {test_area_path}")
                return
        try:
            distance_threshold_text = self.src_distance_edit.text().strip()
            distance_threshold = float(distance_threshold_text or self.src_distance_threshold or 4)
            src_prediction_stride = _optional_positive_int(self.src_prediction_stride_edit.text(), "SRC预测步长")
        except ValueError:
            QMessageBox.critical(self, "参数错误", "矿点命中距离需要是数值，SRC预测步长需要是正整数或留空")
            return
        self._run_async(
            run_dec_task,
            args=(
                models,
                self.log,
                deposit_path,
                {
                    "use_test_area": use_test_area,
                    "test_area_file": test_area_path,
                    "distance_threshold": distance_threshold,
                    "distance_threshold_override": distance_threshold,
                    "probability_mode": self.src_probability_mode_combo.currentData() or PROBABILITY_MODE_CENTER,
                    "prediction_patch_stride": src_prediction_stride,
                },
            ),
            success_message="SRC曲线生成完成",
            result_handler=self._handle_dec_result,
        )

    def _generate_pa_curves(self):
        models = self._get_selected_pa_models()
        if not models:
            QMessageBox.warning(self, "未选择模型", "请至少选择一个模型进行P-A计算")
            return
        name_counts = {}
        for model in models:
            name = str(model.get("name") or "")
            name_counts[name] = name_counts.get(name, 0) + 1
        duplicate_names = {name for name, count in name_counts.items() if count > 1}
        for model in models:
            model["plot_name"] = self._display_pa_model_name(model, duplicate_names)

        h5_path = self.pa_h5_edit["edit"].text().strip()
        mineral_path = self.pa_deposit_edit["edit"].text().strip()
        use_test_area = self.pa_use_test_area_check.isChecked()
        test_area_path = self.pa_test_area_edit["edit"].text().strip()

        for model in models:
            if str(model.get("source_type") or "").lower() != "trained_model":
                continue
            if h5_path and (not model.get("data_path") or not os.path.exists(str(model.get("data_path")))):
                model["data_path"] = h5_path
            if test_area_path and (not model.get("test_area_path") or not os.path.exists(str(model.get("test_area_path")))):
                model["test_area_path"] = test_area_path
            if mineral_path and (
                not use_test_area
                or not model.get("test_mineral_path")
                or not os.path.exists(str(model.get("test_mineral_path")))
            ):
                model["test_mineral_path"] = mineral_path

            if not model.get("data_path"):
                QMessageBox.critical(self, "缺少参数", f"{model.get('name')} 缺少完整H5，请在P-A参数中选择完整H5文件")
                return
            if not os.path.exists(str(model.get("data_path"))):
                QMessageBox.critical(self, "路径不存在", f"完整H5不存在: {model.get('data_path')}")
                return
            if use_test_area and not model.get("test_area_path"):
                QMessageBox.critical(self, "缺少参数", f"{model.get('name')} 缺少测试区域文件，请在P-A参数中选择测试区域文件")
                return
            if use_test_area and not os.path.exists(str(model.get("test_area_path"))):
                QMessageBox.critical(self, "路径不存在", f"测试区域文件不存在: {model.get('test_area_path')}")
                return
            if not model.get("test_mineral_path"):
                QMessageBox.critical(self, "缺少参数", f"{model.get('name')} 缺少矿点文件，请在P-A参数中选择矿点文件")
                return
            if not os.path.exists(str(model.get("test_mineral_path"))):
                QMessageBox.critical(self, "路径不存在", f"矿点文件不存在: {model.get('test_mineral_path')}")
                return

        raw_prediction_models = [
            model
            for model in models
            if str(model.get("source_type") or "").lower() in {"dat", "prediction_h5"}
        ]
        if raw_prediction_models:
            if not mineral_path:
                QMessageBox.critical(self, "缺少参数", "所选 DAT/H5 预测结果需要矿点文件参与计算，请先选择矿点文件")
                return
            if not os.path.exists(mineral_path):
                QMessageBox.critical(self, "路径不存在", f"矿点文件不存在: {mineral_path}")
                return

        if use_test_area:
            trained_models = [
                model
                for model in models
                if str(model.get("source_type") or "").lower() == "trained_model"
            ]
            trained_have_area = all(str(model.get("test_area_path") or "").strip() for model in trained_models)
            if not test_area_path and (raw_prediction_models or not trained_have_area):
                QMessageBox.critical(self, "缺少参数", "独立测试集P-A需要测试区域文件")
                return
            if test_area_path and not os.path.exists(test_area_path):
                QMessageBox.critical(self, "路径不存在", f"测试区域文件不存在: {test_area_path}")
                return

        try:
            distance_threshold = float(self.pa_distance_edit.text() or 4)
            threshold_step = float(self.pa_threshold_step_edit.text() or 0.01)
            pa_prediction_stride = _optional_positive_int(self.pa_prediction_stride_edit.text(), "P-A预测步长")
        except ValueError:
            QMessageBox.critical(self, "参数错误", "距离阈值和阈值扫描步长需要是数值，P-A预测步长需要是正整数或留空")
            return
        if threshold_step <= 0 or threshold_step > 1:
            QMessageBox.critical(self, "参数错误", "阈值扫描步长需要在 0-1 之间")
            return

        self._run_async(
            run_pa_task,
            args=(
                models,
                self.log,
                mineral_path,
                {
                    "use_test_area": use_test_area,
                    "test_area_file": test_area_path,
                    "distance_threshold": distance_threshold,
                    "threshold_step": threshold_step,
                    "probability_mode": self.pa_probability_mode_combo.currentData() or PROBABILITY_MODE_CENTER,
                    "prediction_patch_stride": pa_prediction_stride,
                },
            ),
            success_message="P-A曲线生成完成",
            result_handler=self._handle_pa_result,
        )

    def _handle_dec_result(self, metrics_or_curves):
        if metrics_or_curves and "curves" in metrics_or_curves:
            self.dec_last_result = metrics_or_curves["curves"]
            self._update_dec_plot()
        else:
            self._update_metrics_display(metrics_or_curves)
            self._update_confusion_matrix(metrics_or_curves)

    def _handle_pa_result(self, metrics_or_curves):
        if metrics_or_curves and "curves" in metrics_or_curves:
            self.pa_last_result = metrics_or_curves["curves"]
            self._refresh_pa_model_selector()
            self._update_pa_plot()
            self._update_pa_summary_table(self.pa_last_result)

    def _handle_evaluation_result(self, metrics):
        if not metrics:
            self.log("[警告] 未获取到评估结果")
            return
        self._update_metrics_display(metrics)
        self._update_confusion_matrix(metrics)

    def _handle_shap_result(self, result):
        if not result or "output_dir" not in result:
            self.shap_status_label.setText("任务完成，但未返回输出目录，请查看日志。")
            return
        missing = [
            path for path in (result.get("channel_plot"), result.get("summary_plot"))
            if not path or not os.path.exists(path)
        ]
        if missing:
            self.shap_status_label.setText("SHAP未生成完整图像，请查看日志。")
            QMessageBox.warning(
                self,
                "图像缺失",
                "SHAP任务未生成完整图像文件: " + ", ".join(os.path.basename(path) for path in missing if path),
            )
            return
        output_dir = result["output_dir"]
        if result.get("is_repeat_summary"):
            summary_csv = result.get("importance_csv")
            self.shap_status_label.setText(f"多模型SHAP汇总已保存至: {output_dir}")
            if summary_csv:
                self.log(f"[SHAP] 通道重要性汇总表: {summary_csv}")
        else:
            self.shap_status_label.setText(f"结果已保存至: {output_dir}")
        self._update_shap_images(result)

    def _update_shap_images(self, result):
        self.shap_last_result = dict(result or {})
        self._set_image_preview(self.shap_channel_label, result.get("channel_plot"), "channel")
        self._set_image_preview(self.shap_summary_label, result.get("summary_plot"), "summary")

    def _set_image_preview(self, label: QLabel, image_path: str, key: str):
        if not image_path or not os.path.exists(image_path):
            label.setText("图像文件不存在")
            label.setPixmap(QPixmap())
            self.shap_plot_paths[key] = None
            self.shap_pixmaps[key] = None
            return
        pixmap = QPixmap(image_path)
        if pixmap.isNull():
            label.setText("无法加载图像")
            label.setPixmap(QPixmap())
            self.shap_plot_paths[key] = None
            self.shap_pixmaps[key] = None
            return
        scaled = pixmap.scaled(520, 320, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        label.setPixmap(scaled)
        label.setText("")
        self.shap_plot_paths[key] = image_path
        self.shap_pixmaps[key] = pixmap

    def _open_shap_plot(self, plot_type):
        path = self.shap_plot_paths.get(plot_type)
        if not path or not os.path.exists(path):
            QMessageBox.information(self, "提示", "图像尚未生成或文件不存在")
            return
        try:
            image = plt.imread(path)
        except Exception as exc:
            QMessageBox.critical(self, "加载失败", f"无法读取图像: {exc}")
            return
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.imshow(image)
        ax.axis('off')
        fig.tight_layout()
        plt.show()

    def _shap_importance_table_path(self):
        result = self.shap_last_result or {}
        path = str(result.get("importance_csv") or "").strip()
        if path and os.path.exists(path):
            return path
        output_dir = str(result.get("output_dir") or self.shap_output_edit["edit"].text().strip() or "").strip()
        if output_dir:
            for name in ("shap_channel_importance_summary.csv", "shap_channel_importance.csv"):
                candidate = os.path.join(output_dir, name)
                if os.path.exists(candidate):
                    return candidate
        return ""

    def _shap_summary_values_path(self):
        result = self.shap_last_result or {}
        for key in ("all_summary_values_csv", "summary_values_csv"):
            path = str(result.get(key) or "").strip()
            if path and os.path.exists(path):
                return path
        output_dir = str(result.get("output_dir") or self.shap_output_edit["edit"].text().strip() or "").strip()
        if output_dir:
            for name in ("shap_summary_values_all_runs.csv", "shap_summary_values.csv"):
                candidate = os.path.join(output_dir, name)
                if os.path.exists(candidate):
                    return candidate
        return ""

    def _load_shap_importance_table(self):
        path = self._shap_importance_table_path()
        if not path:
            raise FileNotFoundError("未找到SHAP重要性CSV，请先生成SHAP结果。")
        table = pd.read_csv(path)
        if "channel" not in table.columns:
            raise ValueError(f"SHAP重要性CSV缺少 channel 列: {path}")
        table["channel"] = table["channel"].astype(str)
        if "mean" in table.columns:
            value_col = "mean"
        elif "importance" in table.columns:
            value_col = "importance"
        else:
            raise ValueError(f"SHAP重要性CSV缺少 mean/importance 列: {path}")
        table[value_col] = pd.to_numeric(table[value_col], errors="coerce")
        table = table.dropna(subset=[value_col]).copy()
        if "std" in table.columns:
            table["std"] = pd.to_numeric(table["std"], errors="coerce").fillna(0.0)
        if table.empty:
            raise ValueError("SHAP重要性表中没有可绘制的数据。")
        return table.sort_values(value_col, ascending=False), value_col

    def _shap_display_name(self, channel: str) -> str:
        channel = str(channel)
        text = str(self.shap_channel_name_overrides.get(channel, "")).strip()
        return text or channel

    def _edit_shap_plot_annotations(self, plot_type: str):
        if not self.shap_last_result:
            QMessageBox.information(self, "提示", "请先生成SHAP图像。")
            return
        try:
            table, _ = self._load_shap_importance_table()
        except Exception as exc:
            QMessageBox.critical(self, "读取失败", str(exc))
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("编辑SHAP图件注释")
        dialog.resize(760, 560)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("逐行修改需要替换的通道名称；显示名称留空则恢复原名称。保存后会直接重绘当前SHAP图件。"))

        form = QFormLayout()
        importance_xlabel_edit = QLineEdit(str(self.shap_importance_xlabel or "Mean |SHAP|"))
        summary_xlabel_edit = QLineEdit(str(self.shap_summary_xlabel or "SHAP value (impact on model output)"))
        form.addRow("左图横轴标题", importance_xlabel_edit)
        form.addRow("右图横轴标题", summary_xlabel_edit)
        layout.addLayout(form)

        edit_table = QTableWidget(len(table), 2)
        edit_table.setHorizontalHeaderLabels(["原名称", "显示名称"])
        edit_table.verticalHeader().setVisible(False)
        edit_table.horizontalHeader().setStretchLastSection(True)
        for row, channel in enumerate(table["channel"].tolist()):
            original_item = QTableWidgetItem(str(channel))
            original_item.setFlags(original_item.flags() & ~Qt.ItemIsEditable)
            edit_table.setItem(row, 0, original_item)
            display_text = str(self.shap_channel_name_overrides.get(str(channel), "")).strip()
            edit_table.setItem(row, 1, QTableWidgetItem(display_text))
        layout.addWidget(edit_table, 1)

        btn_row = QHBoxLayout()
        reset_btn = QPushButton("清空显示名称")
        ok_btn = QPushButton("保存并重绘")
        cancel_btn = QPushButton("取消")
        btn_row.addWidget(reset_btn)
        btn_row.addStretch(1)
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        def clear_names():
            for row in range(edit_table.rowCount()):
                edit_table.setItem(row, 1, QTableWidgetItem(""))

        reset_btn.clicked.connect(clear_names)
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)

        if dialog.exec_() != QDialog.Accepted:
            return

        overrides = {}
        for row in range(edit_table.rowCount()):
            original = edit_table.item(row, 0).text().strip()
            display_item = edit_table.item(row, 1)
            display = display_item.text().strip() if display_item else ""
            if original and display and display != original:
                overrides[original] = display
        self.shap_channel_name_overrides = overrides
        self.shap_importance_xlabel = importance_xlabel_edit.text().strip() or "Mean |SHAP|"
        self.shap_summary_xlabel = summary_xlabel_edit.text().strip() or "SHAP value (impact on model output)"

        try:
            self._redraw_shap_plots_from_tables()
            self.log("[SHAP] 已按编辑后的图件注释重绘。")
        except Exception as exc:
            QMessageBox.critical(self, "重绘失败", str(exc))

    def _redraw_shap_plots_from_tables(self):
        result = self.shap_last_result or {}
        output_dir = str(result.get("output_dir") or self.shap_output_edit["edit"].text().strip() or BASE_DIR)
        os.makedirs(output_dir, exist_ok=True)

        table, value_col = self._load_shap_importance_table()
        channels = [str(channel) for channel in table["channel"].tolist()]
        display_names = [self._shap_display_name(channel) for channel in channels]
        values = table[value_col].astype(float).to_numpy()
        show_error_bars = bool(self.shap_show_error_bars_check.isChecked())
        xerr = table["std"].astype(float).to_numpy() if show_error_bars and "std" in table.columns else None

        fig_height = max(4.0, min(12.0, 0.35 * len(table) + 1.5))
        plt.figure(figsize=(10, fig_height))
        y = np.arange(len(table))
        plt.barh(y, values, xerr=xerr, alpha=0.75, color="#4e79a7")
        plt.yticks(y, display_names)
        plt.gca().invert_yaxis()
        plt.xlabel(self.shap_importance_xlabel)
        title_suffix = "with std" if xerr is not None else "no std bars"
        model_count = int(table["count"].max()) if "count" in table.columns and len(table) else 1
        title = f"Mean SHAP Channel Importance ({model_count} models, {title_suffix})" if "mean" in table.columns else "SHAP Channel Importance"
        plt.title(title)
        plt.tight_layout()
        channel_plot = str(result.get("channel_plot") or os.path.join(output_dir, "shap_channel_importance_edited.png"))
        plt.savefig(channel_plot, dpi=300, bbox_inches="tight")
        plt.close()

        summary_plot = str(result.get("summary_plot") or os.path.join(output_dir, "shap_summary_plot_edited.png"))
        summary_values_path = self._shap_summary_values_path()
        if summary_values_path:
            import shap

            values_table = pd.read_csv(summary_values_path)
            if "channel" not in values_table.columns:
                raise ValueError(f"SHAP summary明细缺少 channel 列: {summary_values_path}")
            values_table["channel"] = values_table["channel"].astype(str)
            pivot_index = ["model_index", "sample_order"] if "model_index" in values_table.columns else ["sample_order"]
            shap_wide = values_table.pivot_table(index=pivot_index, columns="channel", values="shap_value", aggfunc="mean")
            feature_wide = values_table.pivot_table(index=pivot_index, columns="channel", values="feature_value", aggfunc="mean")
            common_index = shap_wide.index.intersection(feature_wide.index)
            present_order = [channel for channel in channels if channel in shap_wide.columns and channel in feature_wide.columns]
            if present_order:
                shap_values_2d = shap_wide.loc[common_index, present_order].to_numpy(dtype=np.float64)
                feature_values_2d = feature_wide.loc[common_index, present_order].to_numpy(dtype=np.float64)
                finite_mask = np.isfinite(shap_values_2d).all(axis=1) & np.isfinite(feature_values_2d).all(axis=1)
                shap_values_2d = shap_values_2d[finite_mask]
                feature_values_2d = feature_values_2d[finite_mask]
                if shap_values_2d.size:
                    shap.summary_plot(
                        shap_values_2d,
                        feature_values_2d,
                        feature_names=[self._shap_display_name(channel) for channel in present_order],
                        show=False,
                        plot_type="dot",
                        color_bar=True,
                        max_display=len(present_order),
                        sort=False,
                    )
                    ax = plt.gca()
                    for spine in ax.spines.values():
                        spine.set_visible(True)
                        spine.set_linewidth(1.0)
                        spine.set_color("black")
                    ax.set_xlabel(self.shap_summary_xlabel)
                    plt.gcf().set_size_inches(10, 8)
                    plt.savefig(summary_plot, dpi=300, bbox_inches="tight", facecolor="white")
                    plt.close()

        self.shap_last_result["channel_plot"] = channel_plot
        self.shap_last_result["summary_plot"] = summary_plot
        self._update_shap_images(self.shap_last_result)

    def _run_async(self, target, args=(), success_message=None, result_handler=None):
        if self.active_thread and self.active_thread.is_alive():
            QMessageBox.information(self, "任务运行中", "请等待当前任务完成")
            return

        self.status_label.setText("任务运行中...")

        def worker():
            try:
                result = target(*args)
            except Exception as exc:
                err_msg = str(exc)
                self.log(f"[错误] {err_msg}")
                self._invoke_main(lambda: QMessageBox.critical(self, "执行失败", err_msg))
            else:
                if result_handler:
                    self._invoke_main(lambda r=result: result_handler(r))
                if success_message:
                    self._invoke_main(lambda: QMessageBox.information(self, "完成", success_message))
            finally:
                self._invoke_main(lambda: self.status_label.setText("就绪"))

        self.active_thread = threading.Thread(target=worker, daemon=True)
        self.active_thread.start()

    def _invoke_main(self, callback):
        if threading.current_thread() is threading.main_thread():
            callback()
            return
        self.log_queue.put(("call", callback))

    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_queue.put(("log", f"[{timestamp}] {message}"))

    def _process_log_queue(self):
        updated = False
        while True:
            try:
                kind, payload = self.log_queue.get_nowait()
            except queue.Empty:
                break
            if kind == "log":
                self.log_view.appendPlainText(payload)
                updated = True
            elif kind == "info":
                QMessageBox.information(self, "提示", payload)
            elif kind == "error":
                QMessageBox.critical(self, "错误", payload)
            elif kind == "call":
                try:
                    payload()
                except Exception as exc:
                    # Still surface the error while keeping the event loop responsive
                    QMessageBox.critical(self, "回调执行失败", str(exc))
        if updated:
            bar = self.log_view.verticalScrollBar()
            bar.setValue(bar.maximum())

    def _clear_log(self):
        self.log_view.clear()

    def _format_mean_std(self, values, decimals=4):
        numeric_values = []
        for value in values:
            try:
                number = float(value)
            except (TypeError, ValueError):
                continue
            if np.isfinite(number):
                numeric_values.append(number)
        if not numeric_values:
            return ""
        mean_value = float(np.mean(numeric_values))
        std_value = float(np.std(numeric_values, ddof=1)) if len(numeric_values) > 1 else 0.0
        return f"{mean_value:.{decimals}f} ± {std_value:.{decimals}f}"

    def _set_metrics_table(self, headers, rows):
        self.metrics_table.setColumnCount(len(headers))
        self.metrics_table.setHorizontalHeaderLabels(headers)
        self.metrics_table.setRowCount(0)
        for row_values in rows:
            row_idx = self.metrics_table.rowCount()
            self.metrics_table.insertRow(row_idx)
            for col_idx, value in enumerate(row_values):
                self.metrics_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
        self.metrics_table.resizeColumnsToContents()
        if len(headers) > 1:
            self.metrics_table.horizontalHeader().setStretchLastSection(True)

    def _update_batch_metrics_display(self, metrics):
        models = list(metrics.get("models") or [])
        if not models:
            self._update_metrics_display(None)
            return

        headers = ["指标"] + [str(item.get("model_name") or f"model_{idx}") for idx, item in enumerate(models, start=1)] + ["mean ± std"]
        rows = []

        def add_row(label, getter, formatter, *, aggregate=True, decimals=4):
            raw_values = [getter(item) for item in models]
            display_values = [formatter(value, item) for value, item in zip(raw_values, models)]
            summary = self._format_mean_std(raw_values, decimals=decimals) if aggregate else ""
            rows.append([label] + display_values + [summary])

        if any(item.get("status") for item in models):
            add_row("状态", lambda item: 1 if item.get("status", "成功") == "成功" else 0, lambda _value, item: item.get("status", "成功"), aggregate=False)

        if metrics.get("independent_test"):
            add_row("评估模式", lambda item: 1, lambda _value, _item: "独立测试集", aggregate=False)
            add_row("阈值", lambda item: item.get("best_threshold"), lambda value, _item: "" if value is None else f"{float(value):.4f}")
            add_row("测试矿点检出率 SR", lambda item: item.get("gdr"), lambda value, _item: "" if value is None else f"{float(value):.4f} ({float(value) * 100:.2f}%)")
            add_row(
                "命中测试矿点",
                lambda item: float(item.get("hit_deposits", 0) or 0) / float(item.get("total_deposits", 0) or 1),
                lambda _value, item: f"{item.get('hit_deposits', 0)}/{item.get('total_deposits', 0)}",
            )
            add_row("测试区预测面积占比 PA", lambda item: item.get("par"), lambda value, _item: "" if value is None else f"{float(value):.4f} ({float(value) * 100:.2f}%)")
            add_row("EI = SR/PA", lambda item: item.get("gdr_over_pr"), lambda value, _item: "" if value is None else f"{float(value):.4f}")
            add_row(
                "测试区预测单元",
                lambda item: float((item.get("test_metrics") or {}).get("high_potential_count", 0) or 0) / float((item.get("test_metrics") or {}).get("test_area_count", 0) or 1),
                lambda _value, item: f"{(item.get('test_metrics') or {}).get('high_potential_count', 0)}/{(item.get('test_metrics') or {}).get('test_area_count', 0)}",
            )
            if any("mean_min_distance" in (item.get("test_metrics") or {}) for item in models):
                add_row("平均最近距离", lambda item: (item.get("test_metrics") or {}).get("mean_min_distance"), lambda value, _item: "" if value is None else f"{float(value):.2f}", decimals=2)
            if any("median_min_distance" in (item.get("test_metrics") or {}) for item in models):
                add_row("中位最近距离", lambda item: (item.get("test_metrics") or {}).get("median_min_distance"), lambda value, _item: "" if value is None else f"{float(value):.2f}", decimals=2)
        else:
            add_row("GDR", lambda item: item.get("gdr"), lambda value, _item: "" if value is None else f"{float(value):.4f} ({float(value) * 100:.2f}%)")
            add_row(
                "命中矿点",
                lambda item: float(item.get("hit_deposits", 0) or 0) / float(item.get("total_deposits", 0) or 1),
                lambda _value, item: f"{item.get('hit_deposits', 0)}/{item.get('total_deposits', 0)}",
            )
            add_row("GDR/Pr[f(X)=1]", lambda item: item.get("gdr_over_pr"), lambda value, _item: "" if value is None else f"{float(value):.4f}")
            add_row("PAR", lambda item: item.get("par"), lambda value, _item: "" if value is None else f"{float(value):.4f} ({float(value) * 100:.2f}%)")

        if any(item.get("confusion") for item in models):
            for label, key in (("TP", "tp"), ("FP", "fp"), ("TN", "tn"), ("FN", "fn")):
                add_row(label, lambda item, metric_key=key: (item.get("confusion") or {}).get(metric_key), lambda value, _item: "" if value is None else str(int(value)), decimals=1)
        else:
            rows.append(["混淆矩阵"] + ["无标签信息"] * len(models) + [""])

        self._set_metrics_table(headers, rows)

    def _format_group_summary(self, summary, key, decimals=4):
        stat = summary.get(key) or {}
        mean_value = stat.get("mean", "")
        std_value = stat.get("std", "")
        try:
            mean_number = float(mean_value)
            std_number = float(std_value)
        except (TypeError, ValueError):
            return ""
        if not np.isfinite(mean_number) or not np.isfinite(std_number):
            return ""
        return f"{mean_number:.{decimals}f} ± {std_number:.{decimals}f}"

    def _update_grouped_fold_metrics_display(self, metrics):
        groups = list(metrics.get("groups") or [])
        if not groups:
            self._update_metrics_display(None)
            return

        headers = ["指标"] + [
            str(group.get("display_name") or group.get("folder_name") or f"group_{idx}")
            for idx, group in enumerate(groups, start=1)
        ] + ["mean ± std"]
        rows = []

        def add_row(label, getter, *, aggregate_values=None, decimals=4):
            values = [getter(group) for group in groups]
            if aggregate_values is None:
                summary = ""
            else:
                summary = self._format_mean_std(aggregate_values, decimals=decimals)
            rows.append([label] + values + [summary])

        add_row("状态", lambda group: (group.get("summary") or {}).get("status", ""), aggregate_values=None)
        add_row(
            "有效模型",
            lambda group: f"{(group.get('summary') or {}).get('success_count', 0)}/{(group.get('summary') or {}).get('fold_count', 0)}",
            aggregate_values=None,
        )
        add_row(
            "测试矿点检出率 SR",
            lambda group: self._format_group_summary(group.get("summary") or {}, "sr"),
            aggregate_values=[((group.get("summary") or {}).get("sr") or {}).get("mean") for group in groups],
        )
        add_row(
            "测试区预测面积占比 PA",
            lambda group: self._format_group_summary(group.get("summary") or {}, "pa"),
            aggregate_values=[((group.get("summary") or {}).get("pa") or {}).get("mean") for group in groups],
        )
        add_row(
            "EI = SR/PA",
            lambda group: self._format_group_summary(group.get("summary") or {}, "ei"),
            aggregate_values=[((group.get("summary") or {}).get("ei") or {}).get("mean") for group in groups],
        )
        add_row(
            "阈值",
            lambda group: self._format_group_summary(group.get("summary") or {}, "threshold"),
            aggregate_values=[((group.get("summary") or {}).get("threshold") or {}).get("mean") for group in groups],
        )
        add_row(
            "命中测试矿点",
            lambda group: self._format_group_summary(group.get("summary") or {}, "hit_ratio"),
            aggregate_values=[((group.get("summary") or {}).get("hit_ratio") or {}).get("mean") for group in groups],
        )
        add_row(
            "测试区预测单元",
            lambda group: self._format_group_summary(group.get("summary") or {}, "unit_ratio"),
            aggregate_values=[((group.get("summary") or {}).get("unit_ratio") or {}).get("mean") for group in groups],
        )

        if any((group.get("summary") or {}).get("confusion") for group in groups):
            for label, key in (("TP", "tp"), ("FP", "fp"), ("TN", "tn"), ("FN", "fn")):
                add_row(
                    label,
                    lambda group, metric_key=key: str(((group.get("summary") or {}).get("confusion") or {}).get(metric_key, "")),
                    aggregate_values=[((group.get("summary") or {}).get("confusion") or {}).get(key) for group in groups],
                    decimals=1,
                )
        else:
            rows.append(["混淆矩阵"] + ["无标签信息"] * len(groups) + [""])

        self._set_metrics_table(headers, rows)

    def _update_metrics_display(self, metrics):
        if metrics and metrics.get("grouped_folds"):
            self._update_grouped_fold_metrics_display(metrics)
            return
        if metrics and metrics.get("batch"):
            self._update_batch_metrics_display(metrics)
            return

        self.metrics_table.setColumnCount(2)
        self.metrics_table.setHorizontalHeaderLabels(["指标", "数值"])
        self.metrics_table.setRowCount(0)
        if not metrics:
            self.metrics_table.insertRow(0)
            self.metrics_table.setItem(0, 0, QTableWidgetItem("提示"))
            self.metrics_table.setItem(0, 1, QTableWidgetItem("尚未运行评估"))
            return

        if metrics.get("independent_test"):
            test_metrics = metrics.get("test_metrics") or {}
            rows = [
                ("评估模式", "独立测试集"),
                (
                    "阈值",
                    f"{metrics['best_threshold']:.4f} ({'EI 最大' if test_metrics.get('threshold_strategy') == 'max_ei' else '手动'})",
                ),
                ("测试矿点检出率 SR", f"{metrics['gdr']:.4f} ({metrics['gdr']*100:.2f}%)"),
                ("命中测试矿点", f"{metrics['hit_deposits']}/{metrics['total_deposits']}"),
                ("测试区预测面积占比 PA", f"{metrics['par']:.4f} ({metrics['par']*100:.2f}%)"),
                ("EI = SR/PA", f"{metrics['gdr_over_pr']:.4f}"),
                ("测试区预测单元", f"{test_metrics.get('high_potential_count', 0)}/{test_metrics.get('test_area_count', 0)}"),
            ]
            if "mean_min_distance" in test_metrics:
                rows.append(("平均最近距离", f"{test_metrics['mean_min_distance']:.2f}"))
            if "median_min_distance" in test_metrics:
                rows.append(("中位最近距离", f"{test_metrics['median_min_distance']:.2f}"))
        else:
            rows = [
                ("GDR", f"{metrics['gdr']:.4f} ({metrics['gdr']*100:.2f}%)"),
                ("命中矿点", f"{metrics['hit_deposits']}/{metrics['total_deposits']}")
            ]
            rows.extend([
                ("GDR/Pr[f(X)=1]", f"{metrics['gdr_over_pr']:.4f}"),
                ("PAR", f"{metrics['par']:.4f} ({metrics['par']*100:.2f}%)"),
            ])
        if metrics.get("confusion"):
            conf = metrics["confusion"]
            rows.append(("混淆矩阵", f"TP:{conf['tp']} FP:{conf['fp']} TN:{conf['tn']} FN:{conf['fn']}"))
        else:
            rows.append(("混淆矩阵", "无标签信息"))
        rows.insert(0, ("概率计算方式", _probability_mode_label(metrics.get("probability_mode"))))

        for metric, value in rows:
            row_idx = self.metrics_table.rowCount()
            self.metrics_table.insertRow(row_idx)
            self.metrics_table.setItem(row_idx, 0, QTableWidgetItem(metric))
            self.metrics_table.setItem(row_idx, 1, QTableWidgetItem(value))

    def _clear_confusion_colorbar(self):
        if not self.confusion_colorbar:
            return

        colorbar = self.confusion_colorbar
        self.confusion_colorbar = None

        try:
            colorbar.remove()
        except Exception:
            colorbar_ax = getattr(colorbar, "ax", None)
            if colorbar_ax is not None:
                try:
                    if colorbar_ax in self.confusion_fig.axes:
                        self.confusion_fig.delaxes(colorbar_ax)
                except Exception:
                    pass

    def _update_confusion_matrix(self, metrics):
        self._clear_confusion_colorbar()
        if hasattr(self, "confusion_ax"):
            self.confusion_ax.clear()
        if not metrics or not metrics.get("confusion"):
            self.confusion_ax.axis("off")
            message = "尚未运行评估。" if not metrics else "暂无标签数据"
            self.confusion_ax.text(
                0.5,
                0.5,
                message,
                ha="center",
                va="center",
                transform=self.confusion_ax.transAxes,
                fontsize=12,
                color="gray",
            )
            self.confusion_matrix_data = None
            self.confusion_canvas.draw()
            return

        conf = metrics["confusion"]
        matrix = np.array([[conf["tn"], conf["fp"]], [conf["fn"], conf["tp"]]])
        self.confusion_matrix_data = matrix
        im = self.confusion_ax.imshow(matrix, cmap=plt.cm.Blues)
        self.confusion_colorbar = self.confusion_fig.colorbar(
            im, ax=self.confusion_ax, fraction=0.046, pad=0.04
        )
        self.confusion_ax.set_xticks([0, 1])
        self.confusion_ax.set_yticks([0, 1])
        self.confusion_ax.set_xticklabels(["预测为负", "预测为正"])
        self.confusion_ax.set_yticklabels(["真实为负", "真实为正"])
        self.confusion_ax.set_xlabel("预测标签")
        self.confusion_ax.set_ylabel("真实标签")
        self.confusion_ax.set_title("合计混淆矩阵" if metrics.get("batch") else "混淆矩阵")
        label_map = [["TN", "FP"], ["FN", "TP"]]
        for i in range(2):
            for j in range(2):
                self.confusion_ax.text(
                    j,
                    i,
                    f"{label_map[i][j]}\n{int(matrix[i, j]):,}",
                    ha="center",
                    va="center",
                    color="black",
                    fontsize=11,
                    fontweight="bold",
                )
        self.confusion_fig.tight_layout()
        self.confusion_canvas.draw()

    def _clear_dec_plot(self):
        self.dec_ax.clear()
        self.dec_ax.axis("off")
        self.dec_ax.text(
            0.5,
            0.5,
            "尚未生成SRC曲线",
            ha="center",
            va="center",
            fontsize=12,
            color="gray",
        )
        self.dec_canvas.draw()

    def _curve_display_name(self, curve: dict) -> str:
        name = str(curve.get("legend_name") or curve.get("name") or "").strip()
        return name or "未命名曲线"

    def _src_group_display_name(self, curve: dict) -> str:
        name = str(curve.get("group_name") or curve.get("name") or "").strip()
        if not curve.get("group_name") and name.endswith(")") and " (" in name:
            name = name.rsplit(" (", 1)[0].strip()
        return name or self._curve_display_name(curve)

    def _src_uncertainty_enabled(self) -> bool:
        return bool(
            hasattr(self, "src_uncertainty_band_check")
            and self.src_uncertainty_band_check.isChecked()
        )

    def _src_uncertainty_mode(self) -> str:
        if hasattr(self, "src_uncertainty_mode_combo"):
            return str(self.src_uncertainty_mode_combo.currentData() or "std")
        return "std"

    def _src_curve_step_values(self, curve: dict, x_grid: np.ndarray) -> np.ndarray:
        pac = np.asarray(curve.get("pac") or [], dtype=np.float64).reshape(-1)
        src = np.asarray(curve.get("src") or [], dtype=np.float64).reshape(-1)
        valid = np.isfinite(pac) & np.isfinite(src)
        pac = pac[valid]
        src = src[valid]
        if pac.size == 0 or src.size == 0:
            return np.zeros_like(x_grid, dtype=np.float64)

        order = np.argsort(pac)
        pac = pac[order]
        src = src[order]
        if pac[0] > 0.0:
            pac = np.insert(pac, 0, 0.0)
            src = np.insert(src, 0, 0.0)
        if pac[-1] < 1.0 or src[-1] < 1.0:
            pac = np.append(pac, 1.0)
            src = np.append(src, 1.0)

        idx = np.searchsorted(pac, x_grid, side="right") - 1
        idx = np.clip(idx, 0, len(src) - 1)
        return np.clip(src[idx], 0.0, 1.0)

    def _aggregate_src_group(self, group_name: str, curves: list[dict]):
        valid_curves = [
            curve
            for curve in curves
            if curve.get("pac") is not None and curve.get("src") is not None
        ]
        if not valid_curves:
            return None

        x_values = [0.0, 1.0]
        for curve in valid_curves:
            x_values.extend(float(value) for value in curve.get("pac") or [] if np.isfinite(float(value)))
        x_grid = np.asarray(sorted(set(round(max(0.0, min(1.0, value)), 10) for value in x_values)), dtype=np.float64)
        if x_grid.size < 2:
            x_grid = np.asarray([0.0, 1.0], dtype=np.float64)

        stack = np.vstack([self._src_curve_step_values(curve, x_grid) for curve in valid_curves])
        mean_values = np.mean(stack, axis=0)
        if len(valid_curves) > 1:
            std_values = np.std(stack, axis=0, ddof=1)
        else:
            std_values = np.zeros_like(mean_values)

        mode = self._src_uncertainty_mode()
        if mode == "ci95":
            band = 1.96 * std_values / np.sqrt(max(len(valid_curves), 1))
            band_label = "95% CI"
        else:
            band = std_values
            band_label = "±1 SD"

        return {
            "name": group_name,
            "group_name": group_name,
            "pac": x_grid.tolist(),
            "src": np.clip(mean_values, 0.0, 1.0).tolist(),
            "lower": np.clip(mean_values - band, 0.0, 1.0).tolist(),
            "upper": np.clip(mean_values + band, 0.0, 1.0).tolist(),
            "n": len(valid_curves),
            "band_label": band_label,
            "color": valid_curves[0].get("color", "#1f77b4"),
        }

    def _src_curve_groups(self, curves: list[dict]) -> list[tuple[str, list[dict]]]:
        groups: dict[str, list[dict]] = {}
        for curve in curves or []:
            group_name = self._src_group_display_name(curve)
            groups.setdefault(group_name, []).append(curve)
        return list(groups.items())

    def _draw_src_curves_on_axis(self, ax, curves: list[dict]):
        ax.set_xlabel("PAR")
        ax.set_ylabel("SRC")
        ax.grid(True, alpha=0.3)
        ax.plot([0, 1], [0, 1], "k--", alpha=0.4, label="Random")

        if self._src_uncertainty_enabled():
            curve_groups = self._src_curve_groups(curves)
            group_palette = list(plt.rcParams["axes.prop_cycle"].by_key().get("color", []))
            if not group_palette:
                group_palette = [
                    "#1f77b4",
                    "#ff7f0e",
                    "#2ca02c",
                    "#d62728",
                    "#9467bd",
                    "#8c564b",
                    "#e377c2",
                    "#7f7f7f",
                    "#bcbd22",
                    "#17becf",
                ]
            group_colors = {
                group_name: group_palette[index % len(group_palette)]
                for index, (group_name, _group_curves) in enumerate(curve_groups)
            }

            for group_name, group_curves in curve_groups:
                color = group_colors.get(group_name, "#1f77b4")
                if len(group_curves) > 1:
                    aggregate = self._aggregate_src_group(group_name, group_curves)
                    if not aggregate:
                        continue
                    ax.fill_between(
                        aggregate["pac"],
                        aggregate["lower"],
                        aggregate["upper"],
                        step="post",
                        color=color,
                        alpha=0.18,
                        linewidth=0,
                        label=f"{group_name} {aggregate['band_label']}",
                    )
                    ax.plot(
                        aggregate["pac"],
                        aggregate["src"],
                        drawstyle="steps-post",
                        linewidth=2.0,
                        color=color,
                        label=f"{group_name} mean (n={aggregate['n']})",
                    )
                    self._draw_src_corner_connector(ax, aggregate)
                else:
                    curve = group_curves[0]
                    ax.plot(
                        curve["pac"],
                        curve["src"],
                        drawstyle="steps-post",
                        linewidth=1.8,
                        color=color,
                        label=self._curve_display_name(curve),
                    )
                    self._draw_src_corner_connector(ax, curve)
        else:
            for curve in curves:
                ax.plot(
                    curve["pac"],
                    curve["src"],
                    drawstyle="steps-post",
                    linewidth=1.8,
                    color=curve["color"],
                    label=self._curve_display_name(curve),
                )
                self._draw_src_corner_connector(ax, curve)

        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.set_xticks(np.arange(0, 1.1, 0.1))
        ax.set_yticks(np.arange(0, 1.1, 0.1))
        formatter = plt.FuncFormatter(lambda x, pos: f"{x*100:.0f}%")
        ax.xaxis.set_major_formatter(formatter)
        ax.yaxis.set_major_formatter(formatter)
        ax.set_aspect('equal', adjustable='box')
        ax.legend(loc="lower right", fontsize=9)

    def _edit_curve_legends(self, curve_group: str):
        curve_group = str(curve_group or "").lower()
        if curve_group == "src":
            curves = list(self.dec_last_result or [])
            title = "修改SRC图例"
        elif curve_group == "pa":
            curves = list(self.pa_last_result or [])
            title = "修改P-A图例"
        else:
            return

        if not curves:
            QMessageBox.information(self, "提示", "请先生成曲线后再修改图例。")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.resize(620, 360)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("编辑“显示图例”列；留空会恢复为原始名称。"))

        table = QTableWidget(len(curves), 2)
        table.setHorizontalHeaderLabels(["原始名称", "显示图例"])
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(table, 1)

        for row, curve in enumerate(curves):
            original_name = str(curve.get("name") or f"曲线{row + 1}")
            original_item = QTableWidgetItem(original_name)
            original_item.setFlags(original_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row, 0, original_item)
            table.setItem(row, 1, QTableWidgetItem(self._curve_display_name(curve)))
        table.resizeColumnsToContents()
        table.horizontalHeader().setStretchLastSection(True)

        button_row = QHBoxLayout()
        reset_btn = QPushButton("恢复原名")
        ok_btn = QPushButton("确定")
        cancel_btn = QPushButton("取消")
        button_row.addWidget(reset_btn)
        button_row.addStretch(1)
        button_row.addWidget(ok_btn)
        button_row.addWidget(cancel_btn)
        layout.addLayout(button_row)

        def reset_labels():
            for row, curve in enumerate(curves):
                table.setItem(row, 1, QTableWidgetItem(str(curve.get("name") or f"曲线{row + 1}")))

        reset_btn.clicked.connect(reset_labels)
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)

        if dialog.exec_() != QDialog.Accepted:
            return

        for row, curve in enumerate(curves):
            original_name = str(curve.get("name") or f"曲线{row + 1}")
            item = table.item(row, 1)
            display_name = str(item.text() if item else "").strip() or original_name
            if display_name == original_name:
                curve.pop("legend_name", None)
            else:
                curve["legend_name"] = display_name

        if curve_group == "src":
            self._update_dec_plot()
        else:
            self._refresh_pa_model_selector()
            self._update_pa_summary_table(self.pa_last_result or [])
            self._update_pa_plot()
        self.log(f"[图例] 已更新 {len(curves)} 条曲线显示名称")

    def _update_dec_plot(self):
        if not self.dec_last_result:
            self._clear_dec_plot()
            return

        self.dec_ax.clear()
        self.dec_ax.set_title("SRC曲线")
        self._draw_src_curves_on_axis(self.dec_ax, self.dec_last_result)
        self.dec_fig.tight_layout()
        self.dec_canvas.draw()

    def _show_dec_fullscreen(self):
        if not self.dec_last_result:
            QMessageBox.information(self, "提示", "请先生成SRC曲线")
            return

        fig, ax = plt.subplots(figsize=(8, 6))
        self._draw_src_curves_on_axis(ax, self.dec_last_result)
        fig.tight_layout()
        plt.show()

    def _draw_src_corner_connector(self, ax, curve):
        pac_values = list(curve.get("pac") or [])
        src_values = list(curve.get("src") or [])
        if not pac_values or not src_values:
            return
        last_pac = float(pac_values[-1])
        last_src = float(src_values[-1])
        if abs(last_pac - 1.0) < 1e-12 and abs(last_src - 1.0) < 1e-12:
            return
        connector_x = [last_pac]
        connector_y = [last_src]
        if last_pac < 1.0 - 1e-12:
            connector_x.append(1.0)
            connector_y.append(last_src)
        if last_src < 1.0 - 1e-12:
            connector_x.append(1.0)
            connector_y.append(1.0)
        ax.plot(
            connector_x,
            connector_y,
            linestyle="-",
            linewidth=1.8,
            color=curve.get("color", "gray"),
            alpha=1.0,
            label="_nolegend_",
        )

    def _show_confusion_fullscreen(self):
        if self.confusion_matrix_data is None:
            QMessageBox.information(self, "提示", "暂无有效的混淆矩阵可展示")
            return

        matrix = self.confusion_matrix_data
        fig, ax = plt.subplots(figsize=(5, 5))
        im = ax.imshow(matrix, cmap=plt.cm.Blues)
        fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        ax.set_xticks([0, 1])
        ax.set_yticks([0, 1])
        ax.set_xticklabels(["预测为负", "预测为正"])
        ax.set_yticklabels(["真实为负", "真实为正"])
        ax.set_xlabel("预测标签")
        ax.set_ylabel("真实标签")
        ax.set_title("混淆矩阵")
        label_map = [["TN", "FP"], ["FN", "TP"]]
        for i in range(2):
            for j in range(2):
                ax.text(
                    j,
                    i,
                    f"{label_map[i][j]}\n{int(matrix[i, j]):,}",
                    ha="center",
                    va="center",
                    color="black",
                    fontsize=12,
                    fontweight="bold",
                )
        fig.tight_layout()
        plt.show()

    def _save_dec_curve(self):
        if not self.dec_last_result:
            QMessageBox.information(self, "提示", "请先生成SRC曲线")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存SRC曲线",
            os.path.join(BASE_DIR, "src_curve.png"),
            "PNG (*.png);;PDF (*.pdf);;所有文件 (*.*)",
        )
        if not file_path:
            return
        self.dec_fig.savefig(file_path, dpi=300, bbox_inches='tight')
        QMessageBox.information(self, "保存成功", f"已保存至: {file_path}")

    def _clear_pa_plot(self):
        self.pa_fig.clear()
        self.pa_ax = self.pa_fig.add_subplot(111)
        self.pa_ax.axis("off")
        self.pa_ax.text(
            0.5,
            0.5,
            "尚未生成P-A曲线",
            ha="center",
            va="center",
            fontsize=12,
            color="gray",
        )
        self.pa_canvas.draw()

    def _update_pa_summary_table(self, curves):
        self.pa_summary_table.setRowCount(0)
        for curve in curves or []:
            inter = curve.get("intersection") or {}
            row_idx = self.pa_summary_table.rowCount()
            self.pa_summary_table.insertRow(row_idx)
            values = [
                self._curve_display_name(curve),
                "" if inter.get("threshold") is None else f"{float(inter['threshold']):.4f}",
                "" if inter.get("sr") is None else f"{float(inter['sr']):.4f} ({float(inter['sr']) * 100:.2f}%)",
                "" if inter.get("paf") is None else f"{float(inter['paf']):.4f} ({float(inter['paf']) * 100:.2f}%)",
                "" if inter.get("pa_intersection_y") is None else f"{float(inter['pa_intersection_y']):.4f}",
                inter.get("type", ""),
            ]
            for col_idx, value in enumerate(values):
                self.pa_summary_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
        self.pa_summary_table.resizeColumnsToContents()
        self.pa_summary_table.horizontalHeader().setStretchLastSection(True)

    def _refresh_pa_model_selector(self):
        if not hasattr(self, "pa_model_selector_combo"):
            return
        current_name = self.pa_model_selector_combo.currentText()
        self.pa_model_selector_combo.blockSignals(True)
        self.pa_model_selector_combo.clear()
        for index, curve in enumerate(self.pa_last_result or []):
            self.pa_model_selector_combo.addItem(self._curve_display_name(curve), index)
        if current_name:
            match_index = self.pa_model_selector_combo.findText(current_name)
            if match_index >= 0:
                self.pa_model_selector_combo.setCurrentIndex(match_index)
        self.pa_model_selector_combo.blockSignals(False)
        self._update_pa_selector_state()

    def _update_pa_selector_state(self):
        if not hasattr(self, "pa_model_selector_combo") or not hasattr(self, "pa_display_mode_combo"):
            return
        single_mode = self.pa_display_mode_combo.currentData() == "single"
        self.pa_model_selector_combo.setEnabled(single_mode and self.pa_model_selector_combo.count() > 0)

    def _selected_pa_display_curves(self):
        curves = list(self.pa_last_result or [])
        if not curves or not hasattr(self, "pa_display_mode_combo"):
            return curves
        if self.pa_display_mode_combo.currentData() != "single":
            return curves
        index = self.pa_model_selector_combo.currentData() if hasattr(self, "pa_model_selector_combo") else None
        try:
            index = int(index)
        except (TypeError, ValueError):
            index = 0
        if 0 <= index < len(curves):
            return [curves[index]]
        return curves[:1]

    def _draw_single_pa_panel(self, fig, ax, curve, *, compact=False):
        points = sorted(curve.get("curve") or [], key=lambda item: item.get("threshold", 0.0))
        if not points:
            ax.axis("off")
            ax.text(0.5, 0.5, "无可用P-A曲线数据", ha="center", va="center", color="gray")
            return ax

        x = [float(item["threshold"]) for item in points]
        sr = [float(item["sr"]) for item in points]
        paf = [float(item["paf"]) for item in points]
        color = curve.get("color", "#1f77b4")
        title = self._curve_display_name(curve)

        line_sr, = ax.plot(x, sr, color="#d62728", linewidth=2.0, label="矿点检出率 SR")
        ax.set_title(title, fontsize=10 if compact else 11)
        ax.set_xlabel("概率阈值")
        ax.set_ylabel("矿点检出率 SR")
        ax.set_xlim(0, 1)
        ax.set_ylim(-0.02, 1.02)
        ax.grid(True, linestyle="--", linewidth=0.5, alpha=0.35)

        ax_area = ax.twinx()
        line_area, = ax_area.plot(x, paf, color="#2ca02c", linewidth=2.0, label="预测面积占比 PAF")
        ax_area.set_ylabel("预测面积占比 PAF")
        ax_area.set_ylim(1.02, -0.02)

        formatter = plt.FuncFormatter(lambda value, pos: f"{value * 100:.0f}%")
        ax.yaxis.set_major_formatter(formatter)
        ax_area.yaxis.set_major_formatter(formatter)

        inter = curve.get("intersection") or {}
        if inter.get("threshold") is not None and inter.get("sr") is not None:
            inter_x = float(inter["threshold"])
            inter_y = float(inter["sr"])
            ax.scatter(
                [inter_x],
                [inter_y],
                s=38,
                color=color,
                edgecolor="black",
                linewidth=0.5,
                zorder=5,
            )
            if not compact:
                ax.annotate(
                    f"交点\n阈值={inter_x:.3f}\nSR={inter_y:.1%}\nPAF={float(inter.get('paf', 0.0)):.1%}",
                    xy=(inter_x, inter_y),
                    xytext=(8, -32),
                    textcoords="offset points",
                    fontsize=8,
                    arrowprops=dict(arrowstyle="->", linewidth=0.8),
                    bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="gray", alpha=0.85),
                )

        ax.legend(
            handles=[line_area, line_sr],
            labels=["预测面积占比 PAF", "矿点检出率 SR"],
            loc="best",
            fontsize=8,
        )
        return ax

    def _draw_pa_curves(self, fig, curves):
        fig.clear()
        curves = list(curves or [])
        if not curves:
            ax = fig.add_subplot(111)
            ax.axis("off")
            ax.text(
                0.5,
                0.5,
                "尚未生成P-A曲线",
                ha="center",
                va="center",
                fontsize=12,
                color="gray",
            )
            return ax

        if len(curves) == 1:
            fig.set_size_inches(6.6, 4.6, forward=True)
            ax = fig.add_subplot(111)
            self._draw_single_pa_panel(fig, ax, curves[0], compact=False)
            fig.tight_layout()
            return ax

        ncols = min(3, len(curves))
        nrows = int(np.ceil(len(curves) / ncols))
        fig.set_size_inches(max(6.6, 4.6 * ncols), max(4.6, 3.8 * nrows), forward=True)
        axes = fig.subplots(nrows, ncols, squeeze=False)
        last_ax = axes[0][0]
        for index, ax in enumerate(axes.reshape(-1)):
            if index >= len(curves):
                ax.axis("off")
                continue
            last_ax = self._draw_single_pa_panel(fig, ax, curves[index], compact=True)
        fig.suptitle("P-A曲线", fontsize=12)
        fig.tight_layout()
        return last_ax

    def _update_pa_plot(self):
        if not self.pa_last_result:
            self._clear_pa_plot()
            return
        self._update_pa_selector_state()
        self.pa_ax = self._draw_pa_curves(self.pa_fig, self._selected_pa_display_curves())
        self.pa_canvas.draw()

    def _show_pa_fullscreen(self):
        if not self.pa_last_result:
            QMessageBox.information(self, "提示", "请先生成P-A曲线")
            return
        fig = plt.figure(figsize=(8, 6))
        self._draw_pa_curves(fig, self._selected_pa_display_curves())
        plt.show()

    def _save_pa_curve(self):
        if not self.pa_last_result:
            QMessageBox.information(self, "提示", "请先生成P-A曲线")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存P-A曲线",
            os.path.join(BASE_DIR, "pa_curve.png"),
            "PNG (*.png);;PDF (*.pdf);;CSV (*.csv);;所有文件 (*.*)",
        )
        if not file_path:
            return
        suffix = Path(file_path).suffix.lower()
        if suffix == ".csv":
            rows = []
            for curve in self.pa_last_result:
                model_name = self._curve_display_name(curve)
                for point in curve.get("curve") or []:
                    rows.append(
                        {
                            "model": model_name,
                            "threshold": point.get("threshold", ""),
                            "SR": point.get("sr", ""),
                            "PAF": point.get("paf", ""),
                            "EI": point.get("ei", ""),
                            "hit_deposits": point.get("hit_deposits", ""),
                            "total_deposits": point.get("total_deposits", ""),
                            "high_potential_count": point.get("high_potential_count", ""),
                            "area_count": point.get("area_count", ""),
                        }
                    )
            pd.DataFrame(rows).to_csv(file_path, index=False, encoding="utf-8-sig")
        else:
            self.pa_fig.savefig(file_path, dpi=300, bbox_inches="tight")
        QMessageBox.information(self, "保存成功", f"已保存至: {file_path}")

    def closeEvent(self, event):
        if self.active_thread and self.active_thread.is_alive():
            reply = QMessageBox.question(
                self,
                "任务运行中",
                "当前仍有任务执行，确定要退出吗？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply == QMessageBox.No:
                event.ignore()
                return
        event.accept()


def launch_gui():
    app = QApplication.instance()
    owns_app = False
    if app is None:
        app = QApplication(sys.argv)
        owns_app = True
    window = EvaluationWindow()
    window.show()
    if owns_app:
        app.exec_()


def main():
    launch_gui()


if __name__ == "__main__":
    main()

